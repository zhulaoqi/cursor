"""账期净支出：解析 Billing Invoices 列表并导出 Excel。"""

from __future__ import annotations

import asyncio
import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Optional

from .exporter import (
    _PAID_BILLING_STATUSES,
    _billing_month_key,
    _fetch_billing_list_in_ctx,
    _normalize_status_text,
)
from .logger import get
from .models import Account

log = get("billing_ledger")

_AMOUNT_COL_RE = re.compile(r"(?P<amt>\d+(?:\.\d+)?)\s*USD", re.I)
_REFUND_IN_STATUS_RE = re.compile(
    r"refunded\s*\(\s*(?P<amt>\d+(?:\.\d+)?)\s*(?:USD|usd)?\s*\)",
    re.I,
)


@dataclass(frozen=True)
class BillingListRow:
    email: str
    feishu_email: str
    date_text: str
    description_text: str
    billing_month: str
    status_raw: str
    status: str
    list_amount_usd: Optional[Decimal]
    list_amount_raw: str
    refund_in_status_usd: Optional[Decimal]
    invoice_url: str
    row_note: str = ""


@dataclass
class BillingLedgerSummary:
    email: str
    feishu_email: str
    billing_month: str
    amount_total_usd: Decimal
    refund_total_usd: Decimal
    net_spend_usd: Decimal
    row_count: int
    parse_warnings: list[str] = field(default_factory=list)


def parse_amount_usd(text: str) -> Optional[Decimal]:
    """从 Amount 列文案解析 USD 金额，如 `63.96 USD`。"""
    s = (text or "").strip()
    if not s:
        return None
    m = _AMOUNT_COL_RE.search(s)
    if m:
        return Decimal(m.group("amt"))
    m2 = re.search(r"(?P<amt>\d+(?:\.\d+)?)", s)
    if m2:
        return Decimal(m2.group("amt"))
    return None


def parse_refund_amount_from_status(status_raw: str) -> tuple[str, Optional[Decimal]]:
    """解析 Status 列，返回 (normalized_status, refund_amount)。"""
    status = _normalize_status_text(status_raw)
    if status != "refunded":
        return status, None
    m = _REFUND_IN_STATUS_RE.search(status_raw or "")
    if not m:
        return status, None
    return status, Decimal(m.group("amt"))


def rows_from_billing_page_items(
    items: list[dict],
    *,
    email: str,
    feishu_email: str,
    selected_month: str,
) -> list[BillingListRow]:
    """将浏览器抓取的原始 dict 转为 BillingListRow（仅 paid/refunded + 月份匹配）。"""
    month_key = _billing_month_key(selected_month)
    out: list[BillingListRow] = []
    warnings: list[str] = []

    for raw in items or []:
        if not isinstance(raw, dict):
            continue
        status_raw = str(raw.get("status", "") or "")
        status, refund_amt = parse_refund_amount_from_status(status_raw)
        if status not in _PAID_BILLING_STATUSES:
            continue

        date_text = str(raw.get("date", "") or "")
        row_month = _billing_month_key(date_text)
        if month_key and row_month and row_month != month_key:
            continue

        amount_raw = str(raw.get("amountText", "") or "")
        list_amt = parse_amount_usd(amount_raw)
        note_parts: list[str] = []

        if status == "paid":
            if list_amt is None:
                note_parts.append("Paid 行缺少 Amount 列金额")
        else:
            if list_amt is None:
                note_parts.append("Refunded 行缺少 Amount 列（仍尝试扣减退款额）")
            if refund_amt is None:
                note_parts.append("Refunded 行未解析 Status 括号退款额")

        out.append(
            BillingListRow(
                email=email,
                feishu_email=feishu_email,
                date_text=date_text,
                description_text=str(raw.get("description", "") or ""),
                billing_month=row_month or month_key,
                status_raw=status_raw,
                status=status,
                list_amount_usd=list_amt,
                list_amount_raw=amount_raw,
                refund_in_status_usd=refund_amt,
                invoice_url=str(raw.get("url", "") or ""),
                row_note="；".join(note_parts),
            )
        )

    if warnings:
        log.debug(f"[{email}] rows_from_billing_page_items warnings: {warnings}")
    return out


def summarize_account_ledger(
    rows: list[BillingListRow],
    *,
    email: str,
    feishu_email: str,
    billing_month: str,
) -> BillingLedgerSummary:
    """单账号账期汇总：Σ Amount(paid+refunded) − Σ Status 退款额。"""
    amount_total = Decimal("0")
    refund_total = Decimal("0")
    warnings: list[str] = []
    counted = 0

    for r in rows:
        if r.status == "paid":
            if r.list_amount_usd is not None:
                amount_total += r.list_amount_usd
                counted += 1
            elif r.row_note:
                warnings.append(f"{r.date_text}: {r.row_note}")
        elif r.status == "refunded":
            if r.list_amount_usd is not None:
                amount_total += r.list_amount_usd
            else:
                warnings.append(f"{r.date_text}: Refunded 行无 Amount 列金额")
            if r.refund_in_status_usd is not None:
                refund_total += abs(r.refund_in_status_usd)
                counted += 1
            else:
                warnings.append(f"{r.date_text}: 未解析退款额")

    net = amount_total - refund_total
    return BillingLedgerSummary(
        email=email,
        feishu_email=feishu_email,
        billing_month=billing_month,
        amount_total_usd=amount_total,
        refund_total_usd=refund_total,
        net_spend_usd=net,
        row_count=counted,
        parse_warnings=warnings,
    )


def export_billing_ledger_workbook(
    path: Path,
    summaries: list[BillingLedgerSummary],
    detail_rows: list[BillingListRow],
) -> Path:
    """写入双 Sheet Excel：账期净支出汇总 + 账单原始明细。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from .exporter import _apply_xlsx_style, _cell_align

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    styles = _apply_xlsx_style(wb)

    # ── Sheet 1: 汇总 ──
    ws_sum = wb.active
    ws_sum.title = "账期净支出汇总"
    sum_headers = [
        "账号邮箱",
        "飞书邮箱",
        "账期月份",
        "Amount列合计(USD)",
        "Status退款合计(USD)",
        "账期真实总支出(USD)",
        "参与计算行数",
        "解析备注",
    ]
    ws_sum.append(sum_headers)
    for col in range(1, len(sum_headers) + 1):
        c = ws_sum.cell(row=1, column=col)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]
        c.border = styles["header_border"]

    for s in summaries:
        ws_sum.append([
            s.email,
            s.feishu_email,
            s.billing_month,
            float(s.amount_total_usd),
            float(s.refund_total_usd),
            float(s.net_spend_usd),
            s.row_count,
            "；".join(s.parse_warnings[:5]),
        ])

    for row_idx in range(2, ws_sum.max_row + 1):
        for col_idx in range(1, len(sum_headers) + 1):
            cell = ws_sum.cell(row=row_idx, column=col_idx)
            cell.font = styles["cell_font"]
            cell.fill = styles["row_fill"]
            cell.border = styles["row_border"]
            cell.alignment = _cell_align(cell.value, styles)
        ws_sum.row_dimensions[row_idx].height = 24

    ws_sum.freeze_panes = "A2"
    for col_idx in range(1, len(sum_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 2: 明细 ──
    ws_det = wb.create_sheet("账单原始明细")
    det_headers = [
        "账号邮箱",
        "账期月份",
        "列表日期",
        "描述",
        "日期所属账期",
        "状态原文",
        "状态",
        "列表金额列(USD)",
        "列表金额原文",
        "状态内退款额(USD)",
        "发票链接",
        "行级备注",
    ]
    ws_det.append(det_headers)
    for col in range(1, len(det_headers) + 1):
        c = ws_det.cell(row=1, column=col)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]
        c.border = styles["header_border"]

    for r in detail_rows:
        ws_det.append([
            r.email,
            r.billing_month,
            r.date_text,
            r.description_text,
            r.billing_month,
            r.status_raw,
            r.status,
            float(r.list_amount_usd) if r.list_amount_usd is not None else "",
            r.list_amount_raw,
            float(r.refund_in_status_usd) if r.refund_in_status_usd is not None else "",
            r.invoice_url,
            r.row_note,
        ])

    for row_idx in range(2, ws_det.max_row + 1):
        for col_idx in range(1, len(det_headers) + 1):
            cell = ws_det.cell(row=row_idx, column=col_idx)
            cell.font = styles["cell_font"]
            cell.fill = styles["row_fill"]
            cell.border = styles["row_border"]
            cell.alignment = _cell_align(cell.value, styles)
        ws_det.row_dimensions[row_idx].height = 24

    ws_det.freeze_panes = "A2"
    for col_idx in range(1, len(det_headers) + 1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(path)
    log.info(f"账期净支出 Excel 已写入 {path}")
    return path


async def _scrape_account_billing_list(
    cookie_val: str,
    invoice_month: str,
    browser,
) -> list[dict]:
    from patchright.async_api import Browser

    assert isinstance(browser, Browser)
    ctx = await browser.new_context()
    try:
        await ctx.add_cookies([{
            "name": "WorkosCursorSessionToken",
            "value": cookie_val,
            "domain": "cursor.com",
            "path": "/",
            "httpOnly": True,
            "secure": True,
        }])
        page = await ctx.new_page()
        try:
            return await _fetch_billing_list_in_ctx(page, invoice_month=invoice_month)
        finally:
            await page.close()
    finally:
        await ctx.close()


def scrape_billing_ledger_batch(
    accounts: list[Account],
    invoice_month: str,
    *,
    manager,
    progress_cb: Optional[Callable[[str, str, str], None]] = None,
) -> tuple[list[BillingLedgerSummary], list[BillingListRow]]:
    """并发抓取多账号账单列表并汇总（单 Chromium，不下载 PDF）。"""
    from .api_client import _split_session_token
    from .config import SETTINGS

    month_key = _billing_month_key(invoice_month)
    if not month_key:
        return [], []

    def _cb(email: str, phase: str, msg: str = "") -> None:
        if progress_cb:
            try:
                progress_cb(email, phase, msg)
            except Exception:
                pass

    max_parallel = max(
        1,
        min(SETTINGS.billing_ledger_concurrency, len(accounts)),
    )
    log.info(f"账期净支出并发: {max_parallel}（共 {len(accounts)} 账号）")

    summaries: list[BillingLedgerSummary] = []
    all_details: list[BillingListRow] = []

    async def _one(acc: Account, browser, sem: asyncio.Semaphore) -> None:
        async with sem:
            _cb(acc.email, "fetching", "获取登录态…")
            try:
                token = await asyncio.to_thread(manager.get_valid_token, acc)
            except Exception as e:
                log.warning(f"[{acc.email}] 获取 token 失败: {e}")
                _cb(acc.email, "error", str(e))
                return

            cookie_val, _ = _split_session_token(token)
            if not cookie_val:
                _cb(acc.email, "error", "cookie 为空")
                return

            _cb(acc.email, "ledger", "解析账单列表…")
            retry_times = max(1, min(SETTINGS.billing_ledger_retry_times, 2))
            retry_backoff = max(0, SETTINGS.billing_ledger_retry_backoff_sec)
            raw_items: list[dict] = []
            last_err: Optional[str] = None

            for attempt in range(1, retry_times + 1):
                try:
                    raw_items = await _scrape_account_billing_list(
                        cookie_val, month_key, browser,
                    )
                except Exception as e:
                    last_err = str(e)
                    log.warning(
                        f"[{acc.email}] 账单页抓取异常 "
                        f"({attempt}/{retry_times}): {e}"
                    )
                    if attempt < retry_times:
                        _cb(
                            acc.email,
                            "ledger",
                            f"页面打开失败，{retry_backoff}s 后重试 ({attempt}/{retry_times})…",
                        )
                        await asyncio.sleep(retry_backoff)
                        continue
                    _cb(acc.email, "error", last_err)
                    return

                break

            rows = rows_from_billing_page_items(
                raw_items,
                email=acc.email,
                feishu_email=acc.feishu_email or "",
                selected_month=month_key,
            )
            summary = summarize_account_ledger(
                rows,
                email=acc.email,
                feishu_email=acc.feishu_email or "",
                billing_month=month_key,
            )
            summaries.append(summary)
            all_details.extend(rows)
            msg = f"净支出 {summary.net_spend_usd} USD（{len(rows)} 行）"
            _cb(acc.email, "done", msg)

    async def _run() -> None:
        from patchright.async_api import async_playwright

        sem = asyncio.Semaphore(max_parallel)
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            try:
                tasks = [
                    asyncio.create_task(_one(acc, browser, sem))
                    for acc in accounts
                ]
                await asyncio.gather(*tasks, return_exceptions=True)
            finally:
                await browser.close()

    try:
        asyncio.run(_run())
    except Exception as e:
        log.exception(f"账期净支出批量抓取失败: {e}")

    summaries.sort(key=lambda s: s.email.lower())
    return summaries, all_details
