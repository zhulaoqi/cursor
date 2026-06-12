"""FastAPI Web 服务 — Cursor Account Manager 二期。

端点：
  GET  /                           → 前端页面 (index.html)
  POST /api/upload                 → 上传 CSV/Excel，返回解析出的账号列表
  POST /api/run                    → 执行拉取任务，SSE 流式推送进度
  GET  /api/stream/{task_id}       → SSE 进度流
  GET  /api/task/{task_id}         → 查询任务快照（刷新后用于重连恢复）
  GET  /api/download/{token}       → 下载汇总 Excel 文件
  GET  /api/download_zip/{task_id} → 打包下载所有账号文件（ZIP）
  GET  /api/status                 → 服务状态
"""

from __future__ import annotations

import asyncio
import csv
import datetime
import io
import json
import os
import re
import secrets
import tempfile
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

import zipfile

import uvicorn
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from . import exporter, fetcher
from .alerting import send_alert
from .bi_sync import run_daily_sync
from .account_store import load_accounts
from .config import SETTINGS, web_fetch_thread_workers
from .logger import get
from .models import Account, AccountSnapshot
from .plan_scraper import SpendingPanelBatchItem, fetch_spending_panels_batch
from .scheduler import _try_lock, run_scheduler_loop
from .spending_refresh import (
    materialize_spending_info_from_error,
    persist_spending_panel,
    run_daily_spending_refresh_scheduled,
)
from .sync_log_store import get_default_sync_log_store
from .token_store import get_default_store

log = get("web")


def _purge_old_tmp(max_age_sec: int = 3600) -> None:
    """删除 Temp 目录下所有 cam_web_* 且修改时间超过 max_age_sec 的目录。"""
    import shutil as _shutil
    try:
        tmp_root = Path(tempfile.gettempdir())
        cutoff = time.time() - max_age_sec
        for d in tmp_root.glob("cam_web_*"):
            if d.is_dir() and d.stat().st_mtime < cutoff:
                _shutil.rmtree(d, ignore_errors=True)
                log.info(f"已清理旧临时目录: {d.name}")
    except Exception as _e:
        log.debug(f"清理临时目录失败（忽略）: {_e}")


def _schedule_cleanup(path: Path, delay_sec: int = 600) -> None:
    """delay_sec 秒后在后台线程删除 path 目录。"""
    import shutil as _shutil

    def _do():
        time.sleep(delay_sec)
        try:
            if path.exists():
                _shutil.rmtree(path, ignore_errors=True)
                log.info(f"已定时清理临时目录: {path.name}")
        except Exception as _e:
            log.debug(f"定时清理失败（忽略）: {_e}")

    t = threading.Thread(target=_do, daemon=True)
    t.start()


# ─── 应用 ────────────────────────────────────────────────────────
app = FastAPI(title="Cursor Account Manager", version="2.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_STATIC = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=_STATIC), name="static")


# ─── 运行中任务状态 ────────────────────────────────────────────────
_tasks: Dict[str, Dict[str, Any]] = {}
_download_files: Dict[str, Path] = {}
_task_lock = threading.Lock()
_sync_runtime: Dict[str, Dict[str, Any]] = {}
_scheduler_started = False

_spending_progress_lock = threading.Lock()
_spending_refresh_progress: Dict[str, Any] = {
    "running": False,
    "total": 0,
    "done": 0,
    "current_email": "",
    "ok": 0,
    "failed": 0,
    "phase": "idle",
    "scope": "",
    "started_at": 0,
    "updated_at": 0,
    "message": "",
}

IMAP_HOST_DEFAULT = SETTINGS.default_imap_host
IMAP_PORT_DEFAULT = SETTINGS.default_imap_port
BEIJING_TZ = datetime.timezone(datetime.timedelta(hours=8))


@app.on_event("startup")
async def _start_embedded_scheduler() -> None:
    """Web 服务启动时同步启动 BI 调度循环。"""
    global _scheduler_started
    if _scheduler_started:
        return
    if not SETTINGS.bi_sync_enable and not SETTINGS.spending_refresh_enable:
        log.info("BI 与套餐/按量定时刷新均未启用，跳过调度器启动")
        return
    _scheduler_started = True
    t = threading.Thread(
        target=run_scheduler_loop,
        kwargs={"poll_interval_sec": 30},
        name="cam-bi-scheduler",
        daemon=True,
    )
    t.start()
    log.info(
        "调度器已随 Web 服务启动 bi_cron=%s spending_cron=%s spending_alert=%s",
        SETTINGS.bi_sync_cron,
        SETTINGS.spending_refresh_cron,
        SETTINGS.spending_refresh_alert_enable,
    )


# ─── 数据模型 ─────────────────────────────────────────────────────

class AccountRow(BaseModel):
    email: str
    imap_password: str
    feishu_email: str
    imap_host: Optional[str] = None
    imap_port: Optional[int] = None


class RunRequest(BaseModel):
    accounts: List[AccountRow]
    month: Optional[str] = None        # "YYYY-MM" 账单月份：发票过滤 + 文件命名
    date_from: Optional[str] = None    # "YYYY-MM-DD" 使用明细起始日（含），转 Unix 秒
    date_to: Optional[str] = None      # "YYYY-MM-DD" 使用明细结束日（含）
    with_invoices: bool = True
    with_summary: bool = True
    with_raw: bool = False
    with_billing_ledger: bool = False   # 账期净支出（独立任务，与 PDF 并存）
    # 账期净支出导出模式：
    #   "excel_only"   - 仅生成 Excel 下载（默认，兼容旧行为）
    #   "db_only"      - 仅写库，不生成 Excel
    #   "db_and_excel" - 写库 + 生成 Excel 下载
    ledger_export_mode: str = "excel_only"


class SyncRunRequest(BaseModel):
    biz_date: Optional[str] = None
    trigger: str = "manual"
    emails: List[str] = []


class RefreshAccountPlanRequest(BaseModel):
    emails: List[str] = []


# ─── 辅助 ─────────────────────────────────────────────────────────

def _parse_date_range_to_utc_ts(
    date_from: Optional[str],
    date_to: Optional[str],
) -> tuple[Optional[int], Optional[int]]:
    """把前端选择的北京时间自然日转换成 Cursor API 需要的 UTC 秒时间戳。"""
    start_ts: Optional[int] = None
    end_ts: Optional[int] = None
    if date_from:
        start_dt = datetime.datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=BEIJING_TZ)
        start_ts = int(start_dt.timestamp())
    if date_to:
        end_dt = (
            datetime.datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=BEIJING_TZ)
            + datetime.timedelta(days=1, seconds=-1)
        )
        end_ts = int(end_dt.timestamp())
    return start_ts, end_ts


def _fetch_targets_for_run(
    *,
    with_summary: bool,
    with_invoices: bool,
    with_raw: bool,
    with_billing_ledger: bool = False,
) -> tuple[str, ...]:
    """根据导出选项决定本次需要拉取的 API 数据项。"""
    if with_billing_ledger:
        return ()
    if with_raw:
        return fetcher.DEFAULT_WHAT

    targets: list[str] = []
    if with_summary:
        # Token 汇总依赖 usage_events；其它轻量字段保留用于概览/诊断。
        targets.extend(["usage", "plan", "usage_limit", "usage_events", "stripe"])

    # PDF 下载已改为浏览器账单页路径，不依赖 Cursor invoices API。
    # 仅下载 PDF 时仍需要 fetch_one 获取/刷新 token，但无需额外 API 请求。
    return tuple(targets)


def _should_mark_accounts_done_after_export(*, with_invoices: bool) -> bool:
    """没有账单下载阶段时，Web 层需要在导出完成后补账号终态。"""
    return not with_invoices


def _has_zip_output_requested(*, with_summary: bool, with_invoices: bool, with_raw: bool) -> bool:
    """ZIP 里至少会有一种用户请求的导出内容。"""
    return with_summary or with_invoices or with_raw


def _normalize_email(email: str) -> str:
    """统一邮箱：去空白/零宽字符并转小写，避免 CSV/Excel 导入更新不命中。"""
    if not email:
        return ""
    s = str(email).strip().lower()
    # 常见不可见字符：BOM / 零宽空格 / 零宽连接符
    s = re.sub(r"[\u200b\u200c\u200d\ufeff\u2060]", "", s)
    # 邮箱中不应包含空白，出现时直接移除
    s = re.sub(r"\s+", "", s)
    return s


def _normalize_feishu_email(email: str) -> str:
    return _normalize_email(email)


def _validate_required_feishu_email(value: str, *, label: str) -> str:
    feishu_email = _normalize_feishu_email(value)
    if not feishu_email:
        raise ValueError(f"{label} feishu_email 不能为空")
    if "@" not in feishu_email:
        raise ValueError(f"{label} feishu_email 格式不正确")
    return feishu_email


def _parse_csv_bytes(data: bytes) -> List[AccountRow]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = {str(h).strip().lower() for h in (reader.fieldnames or []) if h is not None}
    if "feishu_email" not in headers and "飞书邮箱" not in headers:
        raise ValueError("CSV 缺少必需列: feishu_email")
    rows: List[AccountRow] = []
    for line_no, r in enumerate(reader, start=2):
        # CSV 列名兼容：大小写/前后空白
        d = {(str(k).strip().lower() if k is not None else ""): (v or "") for k, v in r.items()}
        email = _normalize_email(d.get("email", ""))
        pw = (
            d.get("imap_password")
            or d.get("imap_pwd")
            or d.get("password")
            or ""
        ).strip()
        if not email or not pw:
            continue
        feishu_email = _validate_required_feishu_email(
            d.get("feishu_email") or d.get("飞书邮箱") or "",
            label=f"第 {line_no} 行",
        )
        host = (d.get("imap_host") or "").strip() or None
        port_raw = (d.get("imap_port") or "").strip()
        port = int(port_raw) if port_raw.isdigit() else None
        rows.append(
            AccountRow(
                email=email,
                imap_password=pw,
                feishu_email=feishu_email,
                imap_host=host,
                imap_port=port,
            )
        )
    return rows


def _parse_excel_bytes(data: bytes) -> List[AccountRow]:
    from openpyxl import load_workbook
    # data_only=True：读取公式格子的缓存值而非公式本身
    wb = load_workbook(io.BytesIO(data), data_only=True)

    # 尝试所有 sheet，取解析出账号数最多的那个
    best_rows: List[AccountRow] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers: List[str] = []
        rows: List[AccountRow] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # 跳过全空行
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            if not headers:
                headers = [str(c).strip().lower() if c is not None else "" for c in row]
                # 必须包含 email 列才认为是数据表头
                if "email" not in headers:
                    break
                if "feishu_email" not in headers and "飞书邮箱" not in headers:
                    raise ValueError("Excel 缺少必需列: feishu_email")
                continue
            d = dict(zip(headers, row))
            email = _normalize_email(str(d.get("email") or ""))
            pw = str(
                d.get("imap_password") or d.get("imap_pwd") or d.get("password") or ""
            ).strip()
            if not email or not pw or email.lower() in ("none", "email"):
                continue
            feishu_email = _validate_required_feishu_email(
                str(d.get("feishu_email") or d.get("飞书邮箱") or ""),
                label=f"第 {row_idx + 1} 行",
            )
            host_v = d.get("imap_host")
            host = str(host_v).strip() if host_v and str(host_v).strip() not in ("None", "") else None
            port_v = d.get("imap_port")
            port = int(port_v) if port_v and str(port_v).strip().isdigit() else None
            rows.append(
                AccountRow(
                    email=email,
                    imap_password=pw,
                    feishu_email=feishu_email,
                    imap_host=host,
                    imap_port=port,
                )
            )
        if len(rows) > len(best_rows):
            best_rows = rows
        log.debug(f"Excel sheet '{sheet_name}': 解析到 {len(rows)} 个账号")

    log.info(f"Excel 解析完成，共 {len(best_rows)} 个有效账号（{len(wb.sheetnames)} 个 sheet）")
    return best_rows


def _safe_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._@-]+", "_", s)


def _today_bj() -> str:
    return datetime.datetime.now(BEIJING_TZ).strftime("%Y-%m-%d")


def _has_running_sync_task() -> bool:
    return any(t.get("status") == "running" for t in _sync_runtime.values())


def _default_sync_biz_date() -> str:
    return (datetime.datetime.now(BEIJING_TZ) - datetime.timedelta(days=1)).strftime("%Y-%m-%d")


def _build_live_sync_snapshot(run_id: str) -> dict[str, Any]:
    store = get_default_sync_log_store()
    run = store.get_run(run_id)
    if not run:
        return {"run": None, "stages": [], "account_tail": [], "progress": None}
    stages = store.list_stage_logs(run_id)
    accounts = store.list_account_logs(run_id)
    success = sum(1 for a in accounts if a.get("status") == "success")
    failed = sum(1 for a in accounts if a.get("status") == "failed")
    total = int(run.get("account_snapshot_total") or run.get("account_total") or 0)
    done = len(accounts)
    progress = {
        "total": total,
        "done": done,
        "success": success,
        "failed": failed,
        "running": max(0, total - done),
    }
    return {
        "run": run,
        "stages": stages[-20:],
        "account_tail": accounts[-30:],
        "progress": progress,
    }


def _build_runtime_run_fallback(run_id: str) -> Optional[dict[str, Any]]:
    """当 run 尚未落到 SQLite 时，返回内存任务态，避免前端 404。"""
    with _task_lock:
        items = list(_sync_runtime.values())
    for t in items:
        if str(t.get("run_id") or "") != run_id:
            continue
        task_status = str(t.get("status") or "running")
        result = t.get("result") if isinstance(t.get("result"), dict) else {}
        biz_status = str(result.get("status") or "")
        if task_status == "running":
            status = "running"
        elif task_status == "failed":
            status = "failed"
        else:
            status = biz_status or "finished"
        return {
            "run_id": run_id,
            "biz_date": str(result.get("biz_date") or t.get("biz_date") or ""),
            "trigger_type": str(result.get("trigger_type") or "manual"),
            "status": status,
            "started_at": int(t.get("started_at") or int(time.time())),
            "ended_at": int(t.get("finished_at") or 0) or None,
            "duration_sec": int(max(0, int((t.get("finished_at") or time.time())) - int(t.get("started_at") or time.time()))),
            "account_total": int(result.get("account_total") or 0),
            "account_snapshot_total": int(result.get("account_total") or 0),
            "new_account_count": 0,
            "account_success": int(result.get("account_success") or 0),
            "account_failed": int(result.get("account_failed") or 0),
            "event_total": int(result.get("event_total") or 0),
            "ods_rows": int(result.get("ods_rows") or 0),
            "error_summary": str(t.get("error") or result.get("message") or ""),
            "created_at": int(t.get("started_at") or int(time.time())),
            "updated_at": int(t.get("finished_at") or int(time.time())),
        }
    return None


# ─── 路由 ─────────────────────────────────────────────────────────

@app.get("/favicon.ico", include_in_schema=False)
@app.get("/favicon.svg", include_in_schema=False)
async def favicon():
    path = _STATIC / "favicon.svg"
    return FileResponse(path, media_type="image/svg+xml")


@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = _STATIC / "index.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.post("/api/upload")
async def upload_accounts(file: UploadFile = File(...)):
    """解析上传的 CSV / Excel，返回识别到的账号列表（不执行任何拉取）。"""
    data = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            rows = _parse_csv_bytes(data)
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            rows = _parse_excel_bytes(data)
        else:
            # 先尝试 CSV，失败再 Excel
            try:
                rows = _parse_csv_bytes(data)
            except Exception:
                rows = _parse_excel_bytes(data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail="未解析到有效账号，请检查文件格式")

    return {"count": len(rows), "accounts": [r.model_dump() for r in rows]}


@app.get("/api/accounts/template.xlsx")
async def download_accounts_excel_template():
    """下载账号上传 Excel 模板。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "accounts"
    headers = ["email", "imap_password", "imap_host", "imap_port", "feishu_email"]
    example = [
        "cursor183@eclicktech.com.cn",
        "YourImapPassword",
        "imap.feishu.cn",
        993,
        "owner@example.com",
    ]
    ws.append(headers)
    ws.append(example)
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for idx, width in enumerate((32, 24, 24, 10, 32), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cursor_accounts_template.xlsx"'},
    )


_BILLING_MONTH_RE = re.compile(r"^\d{4}-\d{2}$")


@app.post("/api/run")
async def run_task(req: RunRequest):
    """启动拉取任务，返回 task_id，前端用 /api/stream/{task_id} 监听进度。"""
    if not req.accounts:
        raise HTTPException(status_code=400, detail="账号列表为空")

    if req.with_billing_ledger:
        if not req.month or not _BILLING_MONTH_RE.match(req.month.strip()):
            raise HTTPException(
                status_code=400,
                detail="导出账期净支出需选择有效账单月份（YYYY-MM）",
            )

    task_id = secrets.token_hex(8)
    with _task_lock:
        _tasks[task_id] = {
            "status": "pending",
            "total": len(req.accounts),
            "done": 0,
            "ok": [],
            "fail": [],
            "events": [],        # SSE 消息队列（滑动窗口，最多保留 _MAX_EVENTS 条）
            "events_offset": 0,  # 已被裁剪掉的条数，SSE cursor 需加上此偏移
            "started_at": time.time(),
            "finished_at": None,
        }

    def _run_billing_ledger_job() -> None:
        """账期净支出：浏览器抓 Billing 列表，按 ledger_export_mode 写库/生成 Excel。

        ledger_export_mode:
            "excel_only"   - 仅生成 Excel 下载（默认，兼容旧行为）
            "db_only"      - 仅写库，不生成 Excel
            "db_and_excel" - 写库 + 生成 Excel 下载
        """
        from .billing_ledger import export_billing_ledger_workbook, scrape_billing_ledger_batch
        from .token_manager import get_default_manager

        task = _tasks[task_id]
        month = (req.month or "").strip()
        export_mode = (req.ledger_export_mode or "excel_only").strip()
        need_db = export_mode in ("db_only", "db_and_excel")
        need_excel = export_mode in ("excel_only", "db_and_excel")

        accounts = [
            Account(
                email=a.email,
                imap_password=a.imap_password,
                imap_host=a.imap_host or IMAP_HOST_DEFAULT,
                imap_port=a.imap_port or IMAP_PORT_DEFAULT,
                feishu_email=a.feishu_email,
            )
            for a in req.accounts
        ]
        mgr = get_default_manager()
        task_lock = threading.Lock()

        def _export_cb(email: str, phase: str, msg: str = "") -> None:
            if not email:
                _push(task_id, "global_phase", {"phase": phase, "msg": msg})
                return
            if phase in ("done", "warn_done"):
                with task_lock:
                    if phase == "done":
                        task["ok"].append(email)
                    else:
                        task.setdefault("warn", []).append(email)
                    task["done"] += 1
                _push(task_id, "progress", {"email": email, "phase": phase, "msg": msg})
            elif phase == "error":
                with task_lock:
                    task["done"] += 1
                    task["fail"].append({"email": email, "error": msg or "失败"})
                _push(task_id, "progress", {"email": email, "phase": "error", "msg": msg})
            else:
                _push(task_id, "progress", {"email": email, "phase": phase, "msg": msg})

        try:
            summaries, detail_rows = scrape_billing_ledger_batch(
                accounts,
                month,
                manager=mgr,
                progress_cb=_export_cb,
            )

            dl_token: Optional[str] = None
            db_result_msg: str = ""

            # ── 写库 ───────────────────────────────────────────────────
            if need_db and summaries:
                try:
                    from .billing_ledger_store import get_ledger_store
                    _export_cb("", "db_write", "写入净支出汇总到 MySQL…")
                    store = get_ledger_store()
                    store.ensure_tables()
                    n_sum = store.upsert_summaries(summaries)
                    db_result_msg = f"已写库：{n_sum} 条汇总记录"
                    log.info(f"账期净支出写库完成 month={month} {db_result_msg}")
                    _export_cb("", "db_write", db_result_msg)
                except Exception as db_err:
                    log.exception("账期净支出写库失败")
                    _push(task_id, "error", {"msg": f"写库失败: {db_err}"})
                    if not need_excel:
                        # db_only 模式：写库失败即终止
                        raise

            # ── 生成 Excel ────────────────────────────────────────────
            if need_excel:
                _purge_old_tmp(max_age_sec=3600)
                out_dir = Path(tempfile.mkdtemp(prefix="cam_ledger_"))
                xlsx_name = f"账期净支出_{month}.xlsx"
                xlsx_path = out_dir / xlsx_name
                _export_cb("", "summary", "生成账期净支出 Excel…")
                export_billing_ledger_workbook(xlsx_path, summaries, detail_rows)

                if summaries:
                    dl_token = secrets.token_hex(12)
                    with _task_lock:
                        _download_files[dl_token] = xlsx_path

                with task_lock:
                    task["download_token"] = dl_token
                    task["out_dir"] = str(out_dir)
                    task["has_zip"] = False
            else:
                with task_lock:
                    task["download_token"] = None
                    task["out_dir"] = ""
                    task["has_zip"] = False

            _push(task_id, "ready", {
                "download_token": dl_token,
                "has_zip": False,
                "label": month,
                "run_mode": "billing_ledger",
                "ledger_export_mode": export_mode,
                "db_result_msg": db_result_msg,
            })
        except Exception as e:
            log.exception("账期净支出导出失败")
            _push(task_id, "error", {"msg": f"导出失败: {e}"})

        task["status"] = "finished"
        task["finished_at"] = time.time()
        _push(task_id, "finished", {
            "ok": len(task["ok"]),
            "fail": len(task["fail"]),
        })

    def _worker():
        task = _tasks[task_id]
        task["status"] = "running"
        _push(task_id, "start", {"total": len(req.accounts)})

        if req.with_billing_ledger:
            _run_billing_ledger_job()
            return

        # 解析日期范围 → Unix 秒时间戳
        start_ts: Optional[int] = None
        end_ts: Optional[int] = None
        try:
            start_ts, end_ts = _parse_date_range_to_utc_ts(req.date_from, req.date_to)
        except Exception as e:
            log.warning(f"日期解析失败，忽略日期过滤: {e}")

        accounts = [
            Account(
                email=a.email,
                imap_password=a.imap_password,
                imap_host=a.imap_host or IMAP_HOST_DEFAULT,
                imap_port=a.imap_port or IMAP_PORT_DEFAULT,
                feishu_email=a.feishu_email,
            )
            for a in req.accounts
        ]

        import concurrent.futures

        snaps: List[AccountSnapshot] = []
        snap_lock = threading.Lock()

        # fetch 阶段为纯 HTTP；线程数见 web_fetch_thread_workers（可配 WEB_FETCH_MAX_WORKERS）。
        workers = web_fetch_thread_workers(len(accounts))

        # 记录拉取阶段有 warn 的账号，export 完成后保留 warn 标志
        _warn_emails: set[str] = set()

        fetch_targets = _fetch_targets_for_run(
            with_summary=req.with_summary,
            with_invoices=req.with_invoices,
            with_raw=req.with_raw,
            with_billing_ledger=req.with_billing_ledger,
        )

        def _summarize_snapshot_errors(errors: dict[str, str]) -> str:
            """将 snap.errors 压缩成可读提示，优先展示具体错误内容。"""
            if not errors:
                return ""
            parts: list[str] = []
            for key, val in errors.items():
                detail = str(val or "").replace("\n", " ").strip()
                if len(detail) > 160:
                    detail = detail[:157] + "..."
                if detail:
                    parts.append(f"{key}: {detail}")
                else:
                    parts.append(key)
            if len(parts) > 2:
                return "；".join(parts[:2]) + f"；另有 {len(parts) - 2} 项错误"
            return "；".join(parts)

        def _fetch_one(acc: Account):
            _push(task_id, "progress", {"email": acc.email, "phase": "fetching"})
            try:
                snap = fetcher.fetch_one(
                    acc,
                    what=fetch_targets,
                    start_ts=start_ts,
                    end_ts=end_ts,
                )
                has_errors = bool(snap.errors)
                err_keys   = ",".join(snap.errors.keys()) if has_errors else ""
                err_detail = _summarize_snapshot_errors(snap.errors) if has_errors else ""
                with snap_lock:
                    snaps.append(snap)
                    task["done"] += 1
                    if has_errors:
                        _warn_emails.add(acc.email)
                        task["fail"].append({"email": acc.email, "error": err_detail or err_keys})
                if has_errors:
                    for k, v in snap.errors.items():
                        log.warning(f"[{acc.email}] 拉取失败 {k}: {v}")
                    _push(task_id, "progress", {
                        "email": acc.email, "phase": "fetched_warn",
                        "msg": f"拉取部分失败: {err_detail or err_keys}",
                    })
                else:
                    # 拉取完成，等待后续账单下载后再标 done
                    _push(task_id, "progress", {"email": acc.email, "phase": "fetched"})
            except Exception as e:
                with snap_lock:
                    task["done"] += 1
                    task["fail"].append({"email": acc.email, "error": str(e)})
                _push(task_id, "progress", {"email": acc.email, "phase": "error", "msg": str(e)})

        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
            futs = [pool.submit(_fetch_one, acc) for acc in accounts]
            for f in concurrent.futures.as_completed(futs):
                try:
                    f.result()
                except Exception as e:
                    log.exception(f"并发任务异常: {e}")

        # 生成文件
        if snaps:
            label = req.month or time.strftime("%Y-%m")

            # 清理 1 小时前的旧临时目录（防止本次 startup 之后又有旧任务残留）
            _purge_old_tmp(max_age_sec=3600)

            out_dir = Path(tempfile.mkdtemp(prefix="cam_web_"))

            def _export_cb(email: str, phase: str, msg: str = "") -> None:
                """exporter 阶段回调 → 推送 SSE 进度事件（多线程安全）。"""
                if email:
                    if phase == "done":
                        # warn 账号完成后保持 warn_done，ok 账号才算 done
                        is_warn = email in _warn_emails   # set 读取（GIL 保护）
                        if is_warn:
                            _push(task_id, "progress", {
                                "email": email, "phase": "warn_done",
                                "msg": msg or "部分成功",
                            })
                        else:
                            with snap_lock:               # 与 fetch 阶段用同一把锁
                                task["ok"].append(email)
                            _push(task_id, "progress", {
                                "email": email, "phase": "done", "msg": msg,
                            })
                    else:
                        _push(task_id, "progress", {
                            "email": email, "phase": phase, "msg": msg,
                        })
                else:
                    # 全局阶段（生成汇总 / ZIP 就绪）
                    _push(task_id, "global_phase", {"phase": phase, "msg": msg})

            # 限制 export 阶段同时打开的浏览器数，使用账单下载独立并发。
            _browser_sem = threading.Semaphore(
                max(1, SETTINGS.invoice_download_concurrency)
            )
            try:
                exporter.export_per_account(
                    accounts, snaps, out_dir,
                    with_raw=req.with_raw,
                    with_invoices=req.with_invoices,
                    with_detail_xlsx=False,
                    with_full_summary_xlsx=False,
                    with_summary=req.with_summary,
                    invoice_month=req.month or "",
                    start_date=req.date_from or "",
                    end_date=req.date_to or "",
                    progress_cb=_export_cb,
                    browser_semaphore=_browser_sem,
                )
                if _should_mark_accounts_done_after_export(with_invoices=req.with_invoices):
                    for snap in snaps:
                        _export_cb(snap.email, "done", "")
                # export 完成后立即释放大字段内存，防止 300 账号 CSV 文本堆积
                for _s in snaps:
                    _s.usage_csv_text = None  # type: ignore[assignment]
                    _s.usage_events = None    # type: ignore[assignment]

                dl_token: Optional[str] = None
                if req.with_summary:
                    dl_token = secrets.token_hex(12)
                    summary_xlsx = out_dir / "汇总.xlsx"
                    with _task_lock:                      # _download_files 是全局共享 dict
                        _download_files[dl_token] = summary_xlsx
                has_zip = _has_zip_output_requested(
                    with_summary=req.with_summary,
                    with_invoices=req.with_invoices,
                    with_raw=req.with_raw,
                )
                with snap_lock:                           # task 字段写入用 worker 内部锁
                    task["download_token"] = dl_token
                    task["out_dir"] = str(out_dir)
                    task["has_zip"] = has_zip
                _push(task_id, "ready", {
                    "download_token": dl_token,
                    "has_zip": has_zip,
                    "label": label,
                })
            except Exception as e:
                log.exception("生成文件失败")
                _push(task_id, "error", {"msg": f"生成 Excel 失败: {e}"})

        task["status"] = "finished"
        task["finished_at"] = time.time()
        _push(task_id, "finished", {
            "ok": len(task["ok"]),
            "fail": len(task["fail"]),
        })

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    return {"task_id": task_id}


_MAX_EVENTS = 2000  # 滑动窗口上限：超出后裁掉最旧的一半，防止内存无限增长


def _push(task_id: str, event_type: str, data: dict):
    with _task_lock:
        task = _tasks.get(task_id)
        if task is None:
            return
        task["events"].append({"type": event_type, "data": data})
        # 超过上限时裁掉最旧的一半，同时更新偏移量
        if len(task["events"]) > _MAX_EVENTS:
            trim = _MAX_EVENTS // 2
            task["events"] = task["events"][trim:]
            task["events_offset"] = task.get("events_offset", 0) + trim


@app.get("/api/stream/{task_id}")
async def stream_task(task_id: str):
    """SSE 端点，实时推送任务进度。"""
    if task_id not in _tasks:
        raise HTTPException(status_code=404, detail="任务不存在")

    async def event_generator():
        abs_cursor = 0  # 绝对游标（与 events_offset 配合，不受裁剪影响）
        while True:
            with _task_lock:
                task = _tasks.get(task_id, {})
                events = task.get("events", [])
                offset = task.get("events_offset", 0)
                # 转换为 events 列表内的相对下标，防止越界
                rel = max(0, abs_cursor - offset)
                new_events = events[rel:]
                abs_cursor = offset + len(events)  # 推进到当前末尾
                is_done = task.get("status") == "finished"

            for ev in new_events:
                payload = json.dumps(ev["data"], ensure_ascii=False)
                yield f"event: {ev['type']}\ndata: {payload}\n\n"

            if is_done and not new_events:
                break
            await asyncio.sleep(0.4)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/download/{token}")
async def download_file(token: str):
    path = _download_files.get(token)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail="文件不存在或已过期")
    fname = f"cursor_accounts_{time.strftime('%Y%m%d_%H%M')}.xlsx"
    # 下载后 2 小时清理整个任务目录（Excel 的父目录即 out_dir）
    _schedule_cleanup(path.parent.parent, delay_sec=7200)
    return FileResponse(path, filename=fname, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/task/{task_id}")
async def get_task(task_id: str):
    """返回任务快照，前端刷新后用于恢复进度界面。"""
    with _task_lock:
        task = _tasks.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    # 重连时只返回最新 500 条 events，避免大账号批次传输 MB 级 JSON
    all_events = task.get("events", [])
    offset = task.get("events_offset", 0)
    tail_count = min(500, len(all_events))
    tail_events = all_events[-tail_count:] if tail_count else []
    tail_offset = offset + (len(all_events) - tail_count)
    return {
        "status":         task.get("status"),
        "total":          task.get("total", 0),
        "done":           task.get("done", 0),
        "ok":             task.get("ok", []),
        "fail":           task.get("fail", []),
        "download_token": task.get("download_token"),
        "has_zip":        bool(task.get("has_zip") and task.get("out_dir") and Path(task["out_dir"]).exists()),
        "events":         tail_events,
        "events_offset":  tail_offset,  # 前端重连时以此作为初始 abs_cursor
    }


@app.get("/api/download_zip/{task_id}")
async def download_zip(task_id: str):
    """将任务输出目录打包成 ZIP 供下载（汇总.xlsx + 平铺 PDF）。"""
    with _task_lock:
        task = _tasks.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    out_dir = task.get("out_dir")
    if not out_dir or not Path(out_dir).exists():
        raise HTTPException(status_code=404, detail="输出文件不存在或已清理")

    out_path = Path(out_dir)

    def iter_zip():
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # 根目录：汇总.xlsx（各账号 Token 汇总）
            summary_file = out_path / "汇总.xlsx"
            if summary_file.exists():
                zf.write(summary_file, "汇总.xlsx")

            # 根目录平铺 PDF（不再打包账号明细 Excel / 账号子目录）
            used_names: set[str] = set()
            for acc_dir in sorted(out_path.iterdir()):
                if not acc_dir.is_dir():
                    continue
                invoices_dir = acc_dir / "invoices"
                pdf_src = invoices_dir if invoices_dir.exists() else acc_dir
                for pdf in sorted(pdf_src.glob("*.pdf")):
                    target_name = pdf.name
                    if target_name in used_names:
                        stem, suf = pdf.stem, pdf.suffix
                        n = 2
                        while f"{stem}_{n}{suf}" in used_names:
                            n += 1
                        target_name = f"{stem}_{n}{suf}"
                    used_names.add(target_name)
                    zf.write(pdf, target_name)

                raw_json = acc_dir / "raw.json"
                if raw_json.exists():
                    zf.write(raw_json, f"{acc_dir.name}/raw.json")

        buf.seek(0)
        yield buf.read()
        # ZIP 已写入内存，安排 2 小时后清理磁盘文件
        _schedule_cleanup(out_path, delay_sec=7200)

    fname = f"cursor_account_spending_{time.strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(
        iter_zip(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── 账号库 CRUD ───────────────────────────────────────────────────

@app.get("/api/accounts")
async def list_accounts_api(q: str = "", limit: int = 0):
    """返回账号库账号（含 token 状态），支持按邮箱关键词查询。"""
    store = get_default_store()
    accounts = (
        store.search_accounts(q, limit or 30)
        if q.strip() or limit
        else store.list_accounts()
    )
    token_rows = {r.email: r for r in store.list_all()}
    for acc in accounts:
        rec = token_rows.get(acc["email"])
        acc["token_status"] = rec.status if rec else "unknown"
        acc["token_failures"] = rec.consecutive_failures if rec else 0
    return {"accounts": accounts}


def _summarize_account_refresh_errors(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """聚合失败原因，便于前端提示与排查（不截断单条，由前端展示时裁剪）。"""
    from collections import Counter

    c = Counter((r.get("error") or "").strip() for r in results if not r.get("ok"))
    return [{"error": k, "count": v} for k, v in c.most_common(8) if k]


def _on_demand_currently_open_flag(value: object) -> bool:
    """与账号库展示一致：bool True 或整型 1 视为「按需已开」（避免 ``1 is True`` 为假导致漏告警）。"""
    if value is True:
        return True
    if value == 1:
        return True
    if isinstance(value, str) and value.strip().lower() in ("true", "1", "yes"):
        return True
    return False


def _format_on_demand_alert_table(
    entries: list[tuple[str, str]],
    *,
    body_max_chars: int = 12000,
) -> str:
    """按需告警：Markdown 表格，便于飞书卡片扫读。"""
    if not entries:
        return ""
    total = len(entries)
    lines = [
        f"**共 {total} 个账号**",
        "",
        "| # | 账号邮箱 | 飞书邮箱 |",
        "| :---: | :--- | :--- |",
    ]
    nchars = sum(len(s) for s in lines) + max(0, len(lines) - 1)
    shown = 0
    for i, (email, feishu) in enumerate(entries, start=1):
        fei = (feishu or "").strip() or "—"
        row = f"| {i} | {email} | {fei} |"
        if nchars + len(row) + 1 > body_max_chars:
            break
        lines.append(row)
        nchars += len(row) + 1
        shown += 1
    out = "\n".join(lines)
    if shown < total:
        out += (
            f"\n\n… 共 **{total}** 个账号，上表仅展示 **{shown}** 个（已截断以适配飞书长度）；"
            "完整列表请在账号库筛选「按量付费」列查看。"
        )
    return out


def _account_from_store_row(row: dict[str, Any]) -> Account:
    email = _normalize_email(row.get("email", ""))
    return Account(
        email=email,
        imap_password=str(row.get("imap_password") or ""),
        imap_host=str(row.get("imap_host") or IMAP_HOST_DEFAULT),
        imap_port=int(row.get("imap_port") or IMAP_PORT_DEFAULT),
        feishu_email=str(row.get("feishu_email") or "").strip().lower(),
    )


def _spending_row_result_from_batch_item(
    store: Any,
    row: dict[str, Any],
    item: SpendingPanelBatchItem,
) -> tuple[dict[str, Any], bool, Optional[tuple[str, str]], Optional[tuple[str, str]]]:
    """将批量解析结果转为 API 行结果与按需告警条目。"""
    email = _normalize_email(item.email or row.get("email", ""))
    fei = str(row.get("feishu_email") or "").strip()
    if item.error or item.info is None:
        err = item.error or "消费页解析失败"
        inferred_info = materialize_spending_info_from_error(err)
        if inferred_info is not None:
            persist_spending_panel(store, email, inferred_info)
            row_out: dict[str, Any] = {
                "email": email,
                "ok": True,
                "plan_name": inferred_info.plan_name,
                "on_demand_enabled": inferred_info.on_demand_enabled,
                "on_demand_historical": bool(inferred_info.on_demand_historical),
            }
            if inferred_info.plan_snapshot is not None:
                row_out["plan_status"] = inferred_info.plan_snapshot.status
                row_out["plan_amount"] = (
                    str(inferred_info.plan_snapshot.amount)
                    if inferred_info.plan_snapshot.amount is not None else ""
                )
                row_out["plan_error"] = inferred_info.plan_snapshot.error or ""
            return (row_out, True, None, None)
        store.update_account_spending_snapshot(
            email=email,
            plan_name="",
            on_demand_enabled=None,
            on_demand_historical=None,
            spending_error=err,
        )
        return (
            {
                "email": email,
                "ok": False,
                "plan_name": "",
                "on_demand_enabled": None,
                "error": err,
            },
            False,
            None,
            None,
        )

    info = item.info
    persist_spending_panel(store, email, info)
    row_out: dict[str, Any] = {
        "email": email,
        "ok": True,
        "plan_name": info.plan_name,
        "on_demand_enabled": info.on_demand_enabled,
        "on_demand_historical": bool(info.on_demand_historical),
    }
    if info.plan_snapshot is not None:
        row_out["plan_status"] = info.plan_snapshot.status
        row_out["plan_amount"] = (
            str(info.plan_snapshot.amount) if info.plan_snapshot.amount is not None else ""
        )
        row_out["plan_error"] = info.plan_snapshot.error or ""
    open_entry: Optional[tuple[str, str]] = None
    hist_entry: Optional[tuple[str, str]] = None
    if _on_demand_currently_open_flag(info.on_demand_enabled):
        open_entry = (email, fei)
    elif bool(info.on_demand_historical):
        hist_entry = (email, fei)
    return (row_out, True, open_entry, hist_entry)


@app.get("/api/accounts/spending-refresh-busy")
async def spending_refresh_busy_api():
    """探测 Web 端消费页解析文件锁是否被占用（兼容旧前端，仅返回 busy）。"""
    status = await asyncio.to_thread(_spending_refresh_status_sync)
    return {"busy": status["busy"]}


@app.get("/api/accounts/spending-refresh-status")
async def spending_refresh_status_api():
    """消费页解析占用与进度（全库/行内 Web 刷新持锁期间可轮询）。"""
    return await asyncio.to_thread(_spending_refresh_status_sync)


def _spending_refresh_busy_sync() -> dict[str, bool]:
    return {"busy": _spending_refresh_status_sync()["busy"]}


def _spending_refresh_status_sync() -> dict[str, Any]:
    with _try_lock(SETTINGS.spending_refresh_lock_file) as got:
        busy = not got
    progress = _spending_progress_snapshot()
    return {"busy": busy, "progress": progress}


def _spending_progress_snapshot() -> dict[str, Any]:
    with _spending_progress_lock:
        p = dict(_spending_refresh_progress)
    total = int(p.get("total") or 0)
    done = int(p.get("done") or 0)
    p["percent"] = round(done * 100 / total) if total > 0 else 0
    return p


def _spending_progress_start(*, total: int, scope: str) -> None:
    now = int(time.time())
    with _spending_progress_lock:
        _spending_refresh_progress.update(
            {
                "running": True,
                "total": total,
                "done": 0,
                "current_email": "",
                "ok": 0,
                "failed": 0,
                "phase": "running",
                "scope": scope,
                "started_at": now,
                "updated_at": now,
                "message": "正在解析消费页…" if total else "无待解析账号",
            }
        )


def _spending_progress_before_email(email: str, *, index: int, total: int) -> None:
    with _spending_progress_lock:
        _spending_refresh_progress["current_email"] = email
        _spending_refresh_progress["updated_at"] = int(time.time())
        _spending_refresh_progress["message"] = f"正在解析 {index + 1}/{total}"


def _spending_progress_after_email(*, ok: bool) -> None:
    with _spending_progress_lock:
        _spending_refresh_progress["done"] = int(_spending_refresh_progress.get("done") or 0) + 1
        if ok:
            _spending_refresh_progress["ok"] = int(_spending_refresh_progress.get("ok") or 0) + 1
        else:
            _spending_refresh_progress["failed"] = int(_spending_refresh_progress.get("failed") or 0) + 1
        _spending_refresh_progress["updated_at"] = int(time.time())


def _spending_progress_alerting() -> None:
    with _spending_progress_lock:
        _spending_refresh_progress["phase"] = "alerting"
        _spending_refresh_progress["message"] = "汇总按需账号并发送飞书告警…"
        _spending_refresh_progress["updated_at"] = int(time.time())


def _spending_progress_finish() -> None:
    with _spending_progress_lock:
        total = int(_spending_refresh_progress.get("total") or 0)
        done = int(_spending_refresh_progress.get("done") or 0)
        _spending_refresh_progress.update(
            {
                "running": False,
                "phase": "done",
                "done": max(done, total),
                "current_email": "",
                "message": "解析完成",
                "updated_at": int(time.time()),
            }
        )


def _run_refresh_account_spending_locked(req: RefreshAccountPlanRequest) -> dict[str, Any]:
    """同步执行消费页解析（Playwright）；须在线程池中调用以免阻塞 asyncio 事件循环。"""
    with _try_lock(SETTINGS.spending_refresh_lock_file) as lock_ok:
        if not lock_ok:
            raise HTTPException(
                status_code=423,
                detail="消费页解析任务正在进行中（全库/行内或其它浏览器会话），请结束后再试。",
            )
        store = get_default_store()
        requested = {_normalize_email(e) for e in (req.emails or []) if _normalize_email(e)}
        rows = store.list_accounts()
        targets = [r for r in rows if not requested or _normalize_email(r.get("email", "")) in requested]
        work_emails = [
            _normalize_email(row.get("email", ""))
            for row in targets
            if _normalize_email(row.get("email", ""))
        ]
        scope = "web_bulk" if not requested else "web_partial"
        _spending_progress_start(total=len(work_emails), scope=scope)
        results_by_email: dict[str, dict[str, Any]] = {}
        on_demand_alert_open_map: dict[str, tuple[str, str]] = {}
        on_demand_alert_hist_map: dict[str, tuple[str, str]] = {}
        work_rows = [
            row
            for row in targets
            if _normalize_email(row.get("email", ""))
        ]
        row_by_email = {
            _normalize_email(row.get("email", "")): row for row in work_rows
        }
        accounts = [_account_from_store_row(row) for row in work_rows]
        progress_lock = threading.Lock()
        try:

            def _on_account(email: str, index: int, total: int) -> None:
                with progress_lock:
                    _spending_progress_before_email(email, index=index, total=total)

            def _on_result(item: SpendingPanelBatchItem) -> None:
                with progress_lock:
                    row = row_by_email.get(_normalize_email(item.email), {})
                    row_out, row_ok, open_entry, hist_entry = _spending_row_result_from_batch_item(
                        store, row, item,
                    )
                    email_key = _normalize_email(str(row_out.get("email") or item.email or ""))
                    if email_key:
                        results_by_email[email_key] = row_out
                    if open_entry:
                        on_demand_alert_open_map[_normalize_email(open_entry[0])] = open_entry
                    if hist_entry:
                        on_demand_alert_hist_map[_normalize_email(hist_entry[0])] = hist_entry
                    _spending_progress_after_email(ok=row_ok)

            batch_items = fetch_spending_panels_batch(
                accounts,
                silent=False,
                on_account=_on_account,
                on_result=_on_result,
            )
            # 兜底：理论上 on_result 会覆盖全部账号；若某项回调缺失，补做一次持久化。
            for item in batch_items:
                email_key = _normalize_email(item.email)
                if email_key and email_key in results_by_email:
                    continue
                row = row_by_email.get(email_key, {})
                row_out, row_ok, open_entry, hist_entry = _spending_row_result_from_batch_item(
                    store, row, item,
                )
                if email_key:
                    results_by_email[email_key] = row_out
                if open_entry:
                    on_demand_alert_open_map[_normalize_email(open_entry[0])] = open_entry
                if hist_entry:
                    on_demand_alert_hist_map[_normalize_email(hist_entry[0])] = hist_entry
                _spending_progress_after_email(ok=row_ok)

            on_demand_alert_open = sorted(
                on_demand_alert_open_map.values(), key=lambda x: x[0].lower()
            )
            on_demand_alert_hist = sorted(
                on_demand_alert_hist_map.values(), key=lambda x: x[0].lower()
            )
            results = sorted(
                results_by_email.values(), key=lambda r: str(r.get("email") or "").lower()
            )
            if on_demand_alert_open or on_demand_alert_hist:
                parts: list[str] = []
                if on_demand_alert_open:
                    body = _format_on_demand_alert_table(
                        on_demand_alert_open, body_max_chars=12000
                    )
                    parts.append(
                        "**【当前按需已开启】**\n"
                        "以下账号在消费页为按需开启状态，请及时关注。\n\n"
                        + body
                    )
                if on_demand_alert_hist:
                    body = _format_on_demand_alert_table(
                        on_demand_alert_hist, body_max_chars=12000
                    )
                    parts.append(
                        "**【曾有按需消费，当前已关闭】**\n"
                        "以下账号 On-Demand Spending 仍有金额且 Monthly Limit 为 Disabled，"
                        "表示曾产生过按需消费。\n\n"
                        + body
                    )
                combined = "\n\n".join(parts)
                log.info(
                    "消费页按需告警汇总（单条发送）：按需开=%s 曾按需=%s 正文约 %s 字",
                    len(on_demand_alert_open),
                    len(on_demand_alert_hist),
                    len(combined),
                )
                _spending_progress_alerting()
                send_alert(
                    "按量付费（On-demand）相关账号",
                    combined,
                    level="on_demand",
                )
            return {
                "total": len(results),
                "ok": sum(1 for r in results if r["ok"]),
                "failed": sum(1 for r in results if not r["ok"]),
                "results": results,
                "error_summary": _summarize_account_refresh_errors(results),
            }
        finally:
            _spending_progress_finish()


@app.post("/api/accounts/refresh-spending")
async def refresh_account_spending_api(req: RefreshAccountPlanRequest):
    """从 Spending 页解析按需开关、套餐档位名；同页全文可解析时顺带更新 plan_status / plan_amount。

    使用 ``SETTINGS.spending_refresh_lock_file``：同一时刻仅允许一个 Web 解析请求（全库或行内批量）；
    其它 Web 请求返回 423。静默定时全库走 ``run_daily_spending_refresh_silent``，不再占用此文件锁，
    消费页解析使用单 Chromium 批量抓取（``fetch_spending_panels_batch``）。

    解析在默认线程池执行，避免 Playwright 阻塞 asyncio 事件循环（否则刷新页面时其它 API 无法响应）。
    """
    return await asyncio.to_thread(_run_refresh_account_spending_locked, req)


class SaveAccountsRequest(BaseModel):
    accounts: List[AccountRow]
    overwrite: bool = False
    source: str = "upload"


@app.post("/api/accounts/save")
async def save_accounts_api(req: SaveAccountsRequest):
    """将账号持久化到数据库。
    overwrite=False 时跳过已存在的邮箱；overwrite=True 时强制更新。
    返回 saved / skipped 数量及列表。
    """
    from .token_store import get_default_store
    store = get_default_store()
    existing_rows = store.list_accounts()
    # lower(email) -> stored_email（保留原值用于 overwrite 时清理历史大小写记录）
    existing_ci = {_normalize_email(a["email"]): a["email"] for a in existing_rows}

    # 批次内去重：同邮箱（大小写不同）只保留最后一条，行为与 upsert 一致
    latest_by_email: Dict[str, AccountRow] = {}
    for acc in req.accounts:
        norm = _normalize_email(acc.email)
        if not norm:
            continue
        latest_by_email[norm] = AccountRow(
            email=norm,
            imap_password=acc.imap_password,
            feishu_email=acc.feishu_email,
            imap_host=acc.imap_host,
            imap_port=acc.imap_port,
        )

    saved: List[str] = []
    skipped: List[str] = []
    for norm_email, acc in latest_by_email.items():
        try:
            if not _normalize_feishu_email(acc.feishu_email):
                raise ValueError(f"{norm_email} 缺少飞书邮箱")
            feishu_email = _validate_required_feishu_email(acc.feishu_email, label=norm_email)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        existing_email = existing_ci.get(norm_email)
        if existing_email and not req.overwrite:
            skipped.append(norm_email)
            continue
        # 兼容历史脏数据：若仅大小写不同，先删旧记录再写标准化邮箱
        if existing_email and existing_email != norm_email:
            store.delete_account(existing_email)
        store.upsert_account(
            email=norm_email,
            imap_password=acc.imap_password,
            imap_host=acc.imap_host or IMAP_HOST_DEFAULT,
            imap_port=acc.imap_port or IMAP_PORT_DEFAULT,
            feishu_email=feishu_email,
            source=req.source,
        )
        saved.append(norm_email)
        existing_ci[norm_email] = norm_email

    log.info(f"账号库更新：保存 {len(saved)} 个，跳过 {len(skipped)} 个")
    return {"saved": len(saved), "skipped": len(skipped),
            "saved_emails": saved, "skipped_emails": skipped}


@app.delete("/api/accounts/{email:path}")
async def delete_account_api(email: str):
    """从账号库删除指定账号（不影响 tokens 记录）。"""
    from .token_store import get_default_store
    store = get_default_store()
    store.delete_account(email)
    log.info(f"账号库删除：{email}")
    return {"ok": True, "email": email}


class ResetRequest(BaseModel):
    emails: List[str]


@app.post("/api/reset")
async def reset_accounts(req: ResetRequest):
    """重置账号状态：清除 disabled 标记和失败计数，下次运行时重新尝试登录。"""
    from .token_manager import get_default_manager
    mgr = get_default_manager()
    results = []
    for email in req.emails:
        try:
            mgr.store.reset(email)
            results.append({"email": email, "ok": True})
            log.info(f"[{email}] 账号状态已重置")
        except Exception as e:
            results.append({"email": email, "ok": False, "error": str(e)})
    return {"results": results}


@app.get("/api/account_status")
async def account_status(emails: str = ""):
    """查询账号在 tokens.db 里的状态（disabled/active/无记录）。"""
    from .token_manager import get_default_manager
    mgr = get_default_manager()
    email_list = [e.strip() for e in emails.split(",") if e.strip()]
    out = {}
    for email in email_list:
        rec = mgr.store.get(email)
        if rec is None:
            out[email] = {"status": "unknown", "failures": 0}
        else:
            out[email] = {"status": rec.status, "failures": rec.consecutive_failures}
    return out


@app.get("/api/status")
async def api_status():
    with _task_lock:
        running = sum(1 for t in _tasks.values() if t["status"] == "running")
        finished = sum(1 for t in _tasks.values() if t["status"] == "finished")
    return {
        "running": running,
        "finished": finished,
        "total_tasks": len(_tasks),
        "bi_sync_cron": SETTINGS.bi_sync_cron,
        "spending_refresh_enable": SETTINGS.spending_refresh_enable,
        "spending_refresh_cron": SETTINGS.spending_refresh_cron,
    }


# ─── 每日同步监控 API ───────────────────────────────────────────────

@app.get("/api/sync/today")
async def sync_today():
    store = get_default_sync_log_store()
    run = store.get_latest_run(_today_bj())
    if not run:
        return {"run": None}
    stages = store.list_stage_logs(run["run_id"])
    failed_accounts = store.list_account_logs(run["run_id"], status="failed")
    return {"run": run, "stages": stages, "failed_accounts": failed_accounts}


@app.get("/api/sync/runs")
async def sync_runs(limit: int = 30):
    store = get_default_sync_log_store()
    return {"runs": store.list_runs(limit=limit)}


@app.get("/api/sync/running")
async def sync_running():
    """服务内实时运行态（不依赖数据库历史状态）。"""
    with _task_lock:
        running_count = sum(1 for t in _sync_runtime.values() if t.get("status") == "running")
    return {"running": running_count > 0, "running_count": running_count}


@app.get("/api/sync/run/{run_id}")
async def sync_run_detail(run_id: str):
    store = get_default_sync_log_store()
    run = store.get_run(run_id)
    if not run:
        fallback = _build_runtime_run_fallback(run_id)
        if fallback:
            return {
                "run": fallback,
                "stages": [],
                "accounts": [],
            }
        raise HTTPException(status_code=404, detail="run_id 不存在")
    return {
        "run": run,
        "stages": store.list_stage_logs(run_id),
        "accounts": store.list_account_logs(run_id),
    }


@app.delete("/api/sync/run/{run_id}")
async def sync_run_delete(run_id: str):
    """删除单条同步日志（含阶段/账号明细）。"""
    with _task_lock:
        running_same = any(
            str(t.get("run_id") or "") == run_id and str(t.get("status") or "") == "running"
            for t in _sync_runtime.values()
        )
    if running_same:
        raise HTTPException(status_code=409, detail="该任务仍在运行，无法删除")

    store = get_default_sync_log_store()
    deleted = store.delete_run(run_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="run_id 不存在")
    return {"ok": True, "run_id": run_id}


@app.post("/api/spending/run")
async def trigger_spending_run():
    """手动触发“消费页/按量付费”真实调度链路（与 cron 同实现）。"""
    with _task_lock:
        if _has_running_sync_task():
            raise HTTPException(status_code=409, detail="已有调度任务在执行中，请稍后再试")
        sync_task_id = secrets.token_hex(8)
        _sync_runtime[sync_task_id] = {
            "status": "running",
            "result": None,
            "error": "",
            "run_id": "",
            "biz_date": _today_bj(),
            "started_at": int(time.time()),
            "finished_at": None,
        }

    def _worker() -> None:
        try:
            result = run_daily_spending_refresh_scheduled(trigger_type="manual_spending")
            run_id = str(result.get("run_id") or "")
            biz_date = str(result.get("biz_date") or _today_bj())
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "finished"
                _sync_runtime[sync_task_id]["result"] = result
                _sync_runtime[sync_task_id]["run_id"] = run_id
                _sync_runtime[sync_task_id]["biz_date"] = biz_date
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())
        except Exception as e:
            err = f"{type(e).__name__}: {e}"
            send_alert(
                "套餐/按量付费 手动调度异常",
                f"trigger=manual_spending\nerror={err}",
                level="error",
            )
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "failed"
                _sync_runtime[sync_task_id]["error"] = err
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())

    threading.Thread(target=_worker, daemon=True).start()
    return {
        "sync_task_id": sync_task_id,
        "run_id": "",
        "biz_date": _today_bj(),
    }


@app.post("/api/sync/run")
async def sync_run(req: SyncRunRequest):
    trigger_type = (req.trigger or "manual").strip().lower()
    # 约束：scheduler/daily 必须是“昨日 + 全量账号”
    if trigger_type in {"scheduler", "daily"}:
        target_biz_date = _default_sync_biz_date()
        target_emails: Optional[tuple[str, ...]] = None
    elif trigger_type == "manual":
        target_biz_date = req.biz_date or _default_sync_biz_date()
        target_emails = None
    else:
        target_biz_date = req.biz_date or _default_sync_biz_date()
        target_emails = tuple(req.emails) if req.emails else None

    with _task_lock:
        if _has_running_sync_task():
            raise HTTPException(status_code=409, detail="已有同步任务在执行中，请稍后再试")
        sync_task_id = secrets.token_hex(8)
        sync_run_id = f"{target_biz_date.replace('-', '')}_{secrets.token_hex(4)}"
        _sync_runtime[sync_task_id] = {
            "status": "running",
            "result": None,
            "error": "",
            "run_id": sync_run_id,
            "biz_date": target_biz_date,
            "started_at": int(time.time()),
            "finished_at": None,
        }
    # 触发即落主日志：前端刷新列表无需等待后台线程进入业务逻辑
    store = get_default_sync_log_store()
    store.create_run(
        run_id=sync_run_id,
        biz_date=target_biz_date,
        trigger_type=trigger_type or "manual",
        account_total=0,
        account_snapshot_total=0,
        new_account_count=0,
    )

    def _worker() -> None:
        try:
            result = run_daily_sync(
                biz_date=target_biz_date,
                trigger_type=trigger_type or "manual",
                emails=target_emails,
                run_id=sync_run_id,
            )
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "finished"
                _sync_runtime[sync_task_id]["result"] = result
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())
        except Exception as e:
            err = f"{type(e).__name__}: {e}"
            send_alert(
                "BI 同步任务异常",
                f"run_id={sync_run_id}\nbiz_date={target_biz_date}\ntrigger={trigger_type or 'manual'}\nerror={err}",
                level="error",
            )
            try:
                store.add_stage(run_id=sync_run_id, stage="init", status="failed", message=err)
                store.finish_run(
                    run_id=sync_run_id,
                    status="failed",
                    account_success=0,
                    account_failed=0,
                    event_total=0,
                    ods_rows=0,
                    error_summary=err,
                )
            except Exception:
                pass
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "failed"
                _sync_runtime[sync_task_id]["error"] = err
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())

    threading.Thread(target=_worker, daemon=True).start()
    return {
        "sync_task_id": sync_task_id,
        "run_id": sync_run_id,
        "biz_date": target_biz_date,
    }


@app.get("/api/sync/task/{sync_task_id}")
async def sync_task_status(sync_task_id: str):
    with _task_lock:
        task = dict(_sync_runtime.get(sync_task_id) or {})
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    run_id = task.get("run_id")
    if run_id:
        task["live"] = _build_live_sync_snapshot(str(run_id))
    return task


@app.post("/api/sync/retry/{run_id}")
async def sync_retry(run_id: str, biz_date: Optional[str] = None):
    store = get_default_sync_log_store()
    src = store.get_run(run_id)
    target_biz_date = biz_date or (str(src.get("biz_date")) if src else _default_sync_biz_date())
    sync_run_id = f"{str(target_biz_date).replace('-', '')}_{secrets.token_hex(4)}"
    with _task_lock:
        if _has_running_sync_task():
            raise HTTPException(status_code=409, detail="已有同步任务在执行中，请稍后再试")
        sync_task_id = secrets.token_hex(8)
        _sync_runtime[sync_task_id] = {
            "status": "running",
            "result": None,
            "error": "",
            "run_id": sync_run_id,
            "biz_date": target_biz_date,
            "started_at": int(time.time()),
            "finished_at": None,
        }
    store.create_run(
        run_id=sync_run_id,
        biz_date=target_biz_date,
        trigger_type="retry",
        account_total=0,
        account_snapshot_total=0,
        new_account_count=0,
    )

    def _worker() -> None:
        try:
            failed_store = get_default_sync_log_store()
            failed_emails = tuple(failed_store.list_failed_accounts(run_id))
            if not failed_emails:
                result = {"run_id": sync_run_id, "status": "skipped", "message": "无失败账号"}
            else:
                result = run_daily_sync(
                    biz_date=target_biz_date,
                    trigger_type="retry",
                    emails=failed_emails,
                    run_id=sync_run_id,
                )
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "finished"
                _sync_runtime[sync_task_id]["result"] = result
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())
        except Exception as e:
            err = f"{type(e).__name__}: {e}"
            send_alert(
                "BI 补拉任务异常",
                f"run_id={sync_run_id}\nsource_run_id={run_id}\nbiz_date={target_biz_date}\nerror={err}",
                level="error",
            )
            try:
                store.add_stage(run_id=sync_run_id, stage="init", status="failed", message=err)
                store.finish_run(
                    run_id=sync_run_id,
                    status="failed",
                    account_success=0,
                    account_failed=0,
                    event_total=0,
                    ods_rows=0,
                    error_summary=err,
                )
            except Exception:
                pass
            with _task_lock:
                _sync_runtime[sync_task_id]["status"] = "failed"
                _sync_runtime[sync_task_id]["error"] = err
                _sync_runtime[sync_task_id]["finished_at"] = int(time.time())

    threading.Thread(target=_worker, daemon=True).start()
    return {
        "sync_task_id": sync_task_id,
        "run_id": sync_run_id,
        "biz_date": target_biz_date,
    }


def serve(host: str = "0.0.0.0", port: int = 8765, reload: bool = False):
    uvicorn.run("cam.web_server:app", host=host, port=port, reload=reload)
