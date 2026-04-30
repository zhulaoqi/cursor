"""导出 AccountSnapshot → JSON / CSV / XLSX / PDF(发票)。"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import re
from urllib.parse import urlsplit, urlunsplit
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from .config import SETTINGS
from .logger import get
from .models import Account, AccountSnapshot
from .token_manager import TokenManager, get_default_manager

log = get("export")


def _parse_usage_csv(csv_text: str) -> tuple[list[str], list[list[str]]]:
    """解析使用明细 CSV，返回 (headers, rows)。
    headers：原始列名，原样保留。
    rows：每行的字段值列表。
    """
    reader = csv.reader(io.StringIO(csv_text.strip()))
    raw_rows = list(reader)
    if not raw_rows:
        return [], []
    headers = raw_rows[0]
    data_rows = raw_rows[1:]
    return headers, data_rows


# ═══════════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════════

def _fmt_ts_ms(v: Any) -> str:
    """毫秒时间戳 → 'YYYY-MM-DD HH:MM:SS'；非法返回原值。"""
    if v in (None, "", 0):
        return ""
    try:
        ms = int(v)
        if ms > 10**12:  # 毫秒
            return datetime.fromtimestamp(ms / 1000).strftime("%Y-%m-%d %H:%M:%S")
        if ms > 10**9:   # 秒
            return datetime.fromtimestamp(ms).strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        pass
    return str(v)


def _cents_to_usd(v: Any) -> Any:
    try:
        return round(float(v) / 100.0, 4)
    except (TypeError, ValueError):
        return v


def _dig(d: dict, *path, default=None):
    cur: Any = d
    for k in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
        if cur is None:
            return default
    return cur


def _safe_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._@-]+", "_", s)


def export_json(snapshots: Iterable[AccountSnapshot], out_dir: Path) -> list[Path]:
    """每账号写一份 JSON 到 out_dir/{email}.json。"""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for snap in snapshots:
        p = out_dir / f"{_safe_filename(snap.email)}.json"
        p.write_text(
            json.dumps(asdict(snap), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        paths.append(p)
    log.info(f"JSON 已写入 {out_dir}（{len(paths)} 个账号）")
    return paths


def _plan_name(plan: dict) -> str:
    if not isinstance(plan, dict):
        return ""
    for key in ("name", "planName", "plan", "subscription"):
        v = plan.get(key)
        if isinstance(v, str):
            return v
        if isinstance(v, dict):
            for k2 in ("name", "planName"):
                if v.get(k2):
                    return str(v[k2])
    return ""


def _usage_summary(usage: dict) -> dict[str, Any]:
    """从 GetCurrentPeriodUsage 里提取常用字段（结构可能变，尽量多拿）。"""
    if not isinstance(usage, dict):
        return {}
    flat: dict[str, Any] = {}
    for k in ("periodStart", "periodEnd", "startDate", "endDate"):
        if k in usage:
            flat[k] = usage[k]
    for k, v in usage.items():
        if isinstance(v, (int, float, str, bool)):
            flat[k] = v
    return flat


def export_csv(snapshots: list[AccountSnapshot], out_path: Path) -> Path:
    """
    汇总每账号一行：email, plan, period, 各项配额 / 用量 / 错误摘要。
    列是动态的（按 usage / plan 里出现的标量字段并集），但固定几列在前。
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    for snap in snapshots:
        row: dict[str, Any] = {
            "email": snap.email,
            "fetched_at": snap.fetched_at,
            "plan": _plan_name(snap.plan),
            "errors": ",".join(snap.errors.keys()) if snap.errors else "",
        }
        for k, v in _usage_summary(snap.usage).items():
            row[f"usage.{k}"] = v
        # usage_limit 顶层标量
        if isinstance(snap.usage_limit, dict):
            for k, v in snap.usage_limit.items():
                if isinstance(v, (int, float, str, bool)):
                    row[f"limit.{k}"] = v
        # stripe 订阅状态
        if isinstance(snap.stripe, dict):
            for k, v in snap.stripe.items():
                if isinstance(v, (int, float, str, bool)):
                    row[f"stripe.{k}"] = v
        row["invoice_count"] = len(snap.invoices) if snap.invoices else 0
        rows.append(row)

    all_keys: list[str] = []
    seen = set()
    preferred = ["email", "fetched_at", "plan", "invoice_count", "errors"]
    for k in preferred:
        if k not in seen:
            all_keys.append(k); seen.add(k)
    for row in rows:
        for k in row.keys():
            if k not in seen:
                all_keys.append(k); seen.add(k)

    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in all_keys})

    log.info(f"CSV 汇总已写入 {out_path}（{len(rows)} 行，{len(all_keys)} 列）")
    return out_path


# ═══════════════════════════════════════════════════════════════════
# XLSX 导出（多 sheet：概览 / 使用明细 / 发票）
# ═══════════════════════════════════════════════════════════════════

# 账号概览 sheet 固定列顺序
_OVERVIEW_COLUMNS: list[tuple[str, str]] = [
    ("email",                          "邮箱"),
    ("plan",                           "套餐"),
    ("fetched_at_str",                 "抓取时间"),
    ("stripe.membershipType",          "会员类型"),
    ("stripe.subscriptionStatus",      "订阅状态"),
    ("stripe.isYearlyPlan",            "是否年付"),
    ("stripe.trialEligible",           "试用资格"),
    ("stripe.lastPaymentFailed",       "上次付款失败"),
    ("stripe.pendingCancellationDate", "待取消日期"),
    ("stripe.paymentId",               "Stripe Customer"),
    ("usage.billingCycleStart_str",    "计费周期开始"),
    ("usage.billingCycleEnd_str",      "计费周期结束"),
    ("usage.totalSpend_usd",           "已用 (USD)"),
    ("usage.includedSpend_usd",        "套餐内已用 (USD)"),
    ("usage.remaining_usd",            "剩余 (USD)"),
    ("usage.limit_usd",                "套餐额度 (USD)"),
    ("usage.totalPercentUsed",         "总使用率 %"),
    ("usage.autoPercentUsed",          "Auto 使用率 %"),
    ("usage.apiPercentUsed",           "API 使用率 %"),
    ("usage_events_count",             "使用事件数"),
    ("invoice_count",                  "发票数"),
    ("errors",                         "错误字段"),
]

_EVENT_COLUMNS: list[tuple[str, str]] = [
    ("timestamp_str",      "时间"),
    ("model",              "模型"),
    ("kind",               "类型"),
    ("requestsCosts",      "请求数"),
    ("usageBasedCosts",    "按量计费"),
    ("isTokenBasedCall",   "Token计费"),
    ("isChargeable",       "计费"),
    ("isHeadless",         "Headless"),
    ("inputTokens",        "输入Tokens"),
    ("outputTokens",       "输出Tokens"),
    ("cacheWriteTokens",   "缓存写入Tokens"),
    ("cacheReadTokens",    "缓存读取Tokens"),
    ("totalCents_usd",     "总成本(USD)"),
    ("chargedCents_usd",   "扣费(USD)"),
    ("discountPercentOff", "折扣%"),
    ("cursorTokenFee",     "CursorTokenFee"),
    ("owningUser",         "所属用户"),
]

_INVOICE_COLUMNS: list[tuple[str, str]] = [
    ("email",       "邮箱"),
    ("number",      "账单号"),
    ("id",          "账单ID"),
    ("status",      "状态"),
    ("created_str", "账单日期"),
    ("period_str",  "账期"),
    ("total_usd",   "金额(USD)"),
    ("currency",    "币种"),
    ("pdf_file",    "PDF文件"),
    ("pdf_url",     "PDF链接"),
    ("hosted_url",  "在线查看"),
]


def _overview_row(snap: AccountSnapshot) -> dict[str, Any]:
    usage = snap.usage or {}
    pu = usage.get("planUsage") or {}
    stripe = snap.stripe or {}
    row = {
        "email": snap.email,
        "plan": _plan_name(snap.plan) or stripe.get("individualMembershipType") or "",
        "fetched_at_str": _fmt_ts_ms(snap.fetched_at * 1000) if snap.fetched_at else "",
        "stripe.membershipType":        stripe.get("membershipType", ""),
        "stripe.subscriptionStatus":    stripe.get("subscriptionStatus", ""),
        "stripe.isYearlyPlan":          stripe.get("isYearlyPlan", ""),
        "stripe.trialEligible":         stripe.get("trialEligible", ""),
        "stripe.lastPaymentFailed":     stripe.get("lastPaymentFailed", ""),
        "stripe.pendingCancellationDate": stripe.get("pendingCancellationDate", ""),
        "stripe.paymentId":             stripe.get("paymentId", ""),
        "usage.billingCycleStart_str":  _fmt_ts_ms(usage.get("billingCycleStart")),
        "usage.billingCycleEnd_str":    _fmt_ts_ms(usage.get("billingCycleEnd")),
        "usage.totalSpend_usd":         _cents_to_usd(pu.get("totalSpend")),
        "usage.includedSpend_usd":      _cents_to_usd(pu.get("includedSpend")),
        "usage.remaining_usd":          _cents_to_usd(pu.get("remaining")),
        "usage.limit_usd":              _cents_to_usd(pu.get("limit")),
        "usage.totalPercentUsed":       pu.get("totalPercentUsed"),
        "usage.autoPercentUsed":        pu.get("autoPercentUsed"),
        "usage.apiPercentUsed":         pu.get("apiPercentUsed"),
        "usage_events_count":           len(snap.usage_events or []),
        "invoice_count":                len(snap.invoices or []),
        "errors":                       ",".join(sorted(snap.errors.keys())) if snap.errors else "",
    }
    return row


def _event_rows(snap: AccountSnapshot) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ev in snap.usage_events or []:
        if not isinstance(ev, dict):
            continue
        tu = ev.get("tokenUsage") or {}
        rows.append({
            "email":             snap.email,
            "timestamp_str":     _fmt_ts_ms(ev.get("timestamp")),
            "model":             ev.get("model", ""),
            "kind":              ev.get("kind", ""),
            "requestsCosts":     ev.get("requestsCosts"),
            "usageBasedCosts":   ev.get("usageBasedCosts"),
            "isTokenBasedCall":  ev.get("isTokenBasedCall"),
            "isChargeable":      ev.get("isChargeable"),
            "isHeadless":        ev.get("isHeadless"),
            "inputTokens":       tu.get("inputTokens"),
            "outputTokens":      tu.get("outputTokens"),
            "cacheWriteTokens":  tu.get("cacheWriteTokens"),
            "cacheReadTokens":   tu.get("cacheReadTokens"),
            "totalCents_usd":    _cents_to_usd(tu.get("totalCents")),
            "chargedCents_usd":  _cents_to_usd(ev.get("chargedCents")),
            "discountPercentOff": tu.get("discountPercentOff"),
            "cursorTokenFee":    ev.get("cursorTokenFee"),
            "owningUser":        ev.get("owningUser", ""),
        })
    return rows


def _invoice_rows(snap: AccountSnapshot, pdf_files: Optional[dict] = None) -> list[dict[str, Any]]:
    """pdf_files: {inv_number_or_id: filename} 记录实际下载的 PDF 文件名。"""
    rows: list[dict[str, Any]] = []
    for inv in snap.invoices or []:
        if not isinstance(inv, dict):
            continue
        total = inv.get("total") or inv.get("amount_due") or inv.get("amount_paid")
        created = inv.get("created") or inv.get("createdAt")
        created_ms = created * 1000 if isinstance(created, (int, float)) and created < 10**12 else created
        # 账期：优先用 period_start（账单周期开始），比 created 更能代表"哪个月的账单"
        period_start = inv.get("period_start") or inv.get("periodStart")
        period_end   = inv.get("period_end")   or inv.get("periodEnd")
        if period_start:
            ps_ms = period_start * 1000 if isinstance(period_start, (int, float)) and period_start < 10**12 else period_start
            pe_ms = period_end   * 1000 if isinstance(period_end,   (int, float)) and period_end   < 10**12 else period_end
            period_str = _fmt_ts_ms(ps_ms)[:10]  # 只取日期部分
            if pe_ms:
                period_str += f" ~ {_fmt_ts_ms(pe_ms)[:10]}"
        else:
            period_str = ""
        inv_id  = inv.get("id")     or inv.get("invoiceId")    or ""
        inv_num = inv.get("number") or inv.get("invoiceNumber") or ""
        pdf_key = inv_num or str(inv_id)
        pdf_file = (pdf_files or {}).get(pdf_key, "")
        rows.append({
            "email":       snap.email,
            "id":          inv_id,
            "number":      inv_num,
            "status":      inv.get("status", ""),
            "created_str": _fmt_ts_ms(created_ms),
            "period_str":  period_str,
            "total_usd":   _cents_to_usd(total) if isinstance(total, (int, float)) else total,
            "currency":    inv.get("currency", ""),
            "pdf_file":    pdf_file,
            "pdf_url":     inv.get("invoice_pdf") or inv.get("invoicePdf") or inv.get("pdf") or "",
            "hosted_url":  inv.get("hosted_invoice_url") or inv.get("hostedInvoiceUrl") or "",
        })
    return rows


def _apply_xlsx_style(wb_obj) -> dict:
    """返回通用样式对象字典。

    方向一优化（可读性）：
      - 字体 11pt，行高 24px，表头 30px
      - 数字列右对齐，日期列居中，文本左对齐
      - 表头：浅灰底 #EDEEF1，深色加粗，中粗下边框
      - 数据行：纯白底，仅底部细线分隔
      - 合计行：稍深灰底 #DDE0E6，加粗
    """
    import re as _re
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin   = Side(style="thin",   color="D5D8DC")
    medium = Side(style="medium", color="A0A7B4")
    none   = Side(style=None)

    return {
        # 表头
        "header_fill":    PatternFill("solid", fgColor="EDEEF1"),
        "header_font":    Font(bold=True, color="1A1A2E", size=11),
        "header_align":   Alignment(horizontal="center", vertical="center", wrap_text=False),
        "header_border":  Border(bottom=medium, left=none, right=none, top=none),
        # 数据行
        "row_fill":       PatternFill("solid", fgColor="FFFFFF"),
        "row_border":     Border(bottom=thin, left=none, right=none, top=none),
        "cell_font":      Font(size=11, color="2C2C3E"),
        "text_align":     Alignment(horizontal="left",   vertical="center"),
        "num_align":      Alignment(horizontal="right",  vertical="center"),
        "date_align":     Alignment(horizontal="center", vertical="center"),
        # 合计行
        "total_fill":     PatternFill("solid", fgColor="DDE0E6"),
        "total_font":     Font(bold=True, size=11, color="1A1A2E"),
        "total_border":   Border(top=thin, bottom=thin, left=none, right=none),
        # 辅助：日期模式（用于对齐判断）
        "_date_re":       _re.compile(r"^\d{4}-\d{2}"),
    }


def _cell_align(value: Any, styles: dict):
    """根据单元格值类型返回对齐方式。"""
    if isinstance(value, (int, float)):
        return styles["num_align"]
    if isinstance(value, str) and styles["_date_re"].match(value):
        return styles["date_align"]
    return styles["text_align"]


def _style_sheet(ws, n_cols: int, n_rows: int, styles: dict) -> None:
    """对 ws 应用简约商务样式（方向一）。"""
    # 表头行
    ws.row_dimensions[1].height = 30
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill      = styles["header_fill"]
        c.font      = styles["header_font"]
        c.alignment = styles["header_align"]
        c.border    = styles["header_border"]
    # 数据行：高度 + 字体 + 边框 + 值类型对齐
    for row in range(2, n_rows + 2):
        ws.row_dimensions[row].height = 24
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill      = styles["row_fill"]
            c.font      = styles["cell_font"]
            c.border    = styles["row_border"]
            c.alignment = _cell_align(c.value, styles)


def export_xlsx(
    snapshots: list[AccountSnapshot],
    out_path: Path,
    *,
    pdf_files_by_email: Optional[dict[str, dict[str, str]]] = None,
) -> Path:
    """导出多 Sheet Excel：账号概览 / 使用明细。商务配色版。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    styles = _apply_xlsx_style(wb)

    def _write_sheet(ws, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
        keys   = [c[0] for c in columns]
        titles = [c[1] for c in columns]
        ws.append(titles)
        for r in rows:
            ws.append([r.get(k, "") for k in keys])
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{len(rows)+1}"
        # 列宽自适应
        for col_idx, key in enumerate(keys, start=1):
            max_len = len(str(titles[col_idx - 1]))
            for r in rows:
                v = r.get(key, "")
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(max_len + 4, 56))
        _style_sheet(ws, len(keys), len(rows), styles)

    # Sheet 1: 账号概览
    ws1 = wb.active
    ws1.title = "账号概览"
    _write_sheet(ws1, _OVERVIEW_COLUMNS, [_overview_row(s) for s in snapshots])

    # Sheet 2: 使用明细
    ws2 = wb.create_sheet("使用明细")
    _ev_row_count = 0
    csv_texts = [s.usage_csv_text for s in snapshots if s.usage_csv_text]
    if csv_texts:
        all_headers: list[str] = []
        all_data: list[list[str]] = []
        for s in snapshots:
            if not s.usage_csv_text:
                continue
            hdrs, rows = _parse_usage_csv(s.usage_csv_text)
            if not all_headers:
                all_headers = hdrs
            all_data.extend(rows)

        ws2.append(all_headers)
        for row in all_data:
            ws2.append(row)
        ws2.freeze_panes = "A2"
        if all_data:
            ws2.auto_filter.ref = (
                f"A1:{get_column_letter(len(all_headers))}{len(all_data)+1}"
            )
        for col_idx, h in enumerate(all_headers, start=1):
            max_len = len(str(h))
            for row in all_data:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws2.column_dimensions[get_column_letter(col_idx)].width = max(14, min(max_len + 4, 56))
        _style_sheet(ws2, len(all_headers), len(all_data), styles)
        _ev_row_count = len(all_data)
        log.info(f"使用明细 sheet: CSV 格式，{_ev_row_count} 行，{len(all_headers)} 列")
    else:
        event_rows: list[dict[str, Any]] = []
        for s in snapshots:
            event_rows.extend(_event_rows(s))
        _write_sheet(ws2, _EVENT_COLUMNS, event_rows)
        _ev_row_count = len(event_rows)
        log.info(f"使用明细 sheet: API 格式，{_ev_row_count} 行")

    wb.save(out_path)
    log.info(f"XLSX 已写入 {out_path}（账号 {len(snapshots)} / 使用事件 {_ev_row_count}）")
    return out_path


def export_token_summary_xlsx(
    snapshots: list[AccountSnapshot],
    out_path: Path,
    start_date: str = "",
    end_date: str = "",
) -> Path:
    """汇总.xlsx：每个账号一行，各类 Token 合计。

    自动识别 CSV 中含 input/output/cache/token 的列并求和。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # ── 识别 Token 相关列 ──────────────────────────────────────────
    _TOKEN_KW = ("input", "output", "cache", "token")

    def _is_token_col(name: str) -> bool:
        n = name.lower()
        return any(kw in n for kw in _TOKEN_KW)

    def _safe_int(v: str) -> int:
        try:
            return int(str(v).replace(",", "").strip())
        except Exception:
            return 0

    # ── 收集所有账号的 Token 数据 ──────────────────────────────────
    # 先从第一个有 CSV 的 snapshot 取列头
    token_col_names: list[str] = []
    for s in snapshots:
        if s.usage_csv_text:
            hdrs, _ = _parse_usage_csv(s.usage_csv_text)
            token_col_names = [h for h in hdrs if _is_token_col(h)]
            break

    rows_data: list[dict] = []
    for s in snapshots:
        rec: dict[str, Any] = {"账号": s.email, "统计开始": start_date, "统计截止": end_date}
        col_totals: dict[str, int] = {c: 0 for c in token_col_names}

        if s.usage_csv_text:
            hdrs, rows = _parse_usage_csv(s.usage_csv_text)
            col_idx_map = {h: i for i, h in enumerate(hdrs) if _is_token_col(h)}
            for row in rows:
                for col_name, idx in col_idx_map.items():
                    v = _safe_int(row[idx]) if idx < len(row) else 0
                    col_totals[col_name] = col_totals.get(col_name, 0) + v
            rec["记录条数"] = max(0, len(rows))
        else:
            rec["记录条数"] = len(s.usage_events)

        # 构建列名：总 {原列名} 用量
        for col_name in token_col_names:
            rec[f"总{col_name}用量"] = col_totals.get(col_name, 0)
        rows_data.append(rec)

    # ── 合计行 ──────────────────────────────────────────────────────
    if rows_data:
        total_row: dict[str, Any] = {"账号": "合计", "统计开始": "", "统计截止": ""}
        total_row["记录条数"] = sum(r.get("记录条数", 0) for r in rows_data)
        for col_name in token_col_names:
            k = f"总{col_name}用量"
            total_row[k] = sum(r.get(k, 0) for r in rows_data)
        rows_data.append(total_row)

    # ── 写 Excel ────────────────────────────────────────────────────
    wb = Workbook()
    styles = _apply_xlsx_style(wb)
    ws = wb.active
    ws.title = "Token汇总"

    if not rows_data:
        ws.append(["暂无数据"])
        wb.save(out_path)
        return out_path

    # 列顺序：固定前缀 + 各 token 列（Total Tokens 已含合计，不再追加冗余列）
    fixed_cols = ["账号", "统计开始", "统计截止", "记录条数"]
    token_display_cols = [f"总{c}用量" for c in token_col_names]
    all_cols = fixed_cols + token_display_cols

    ws.append(all_cols)
    n_data = len(rows_data) - 1  # 最后一行是合计，单独处理
    for i, r in enumerate(rows_data[:-1]):
        ws.append([r.get(c, "") for c in all_cols])

    # 合计行
    total = rows_data[-1]
    ws.append([total.get(c, "") for c in all_cols])

    ws.freeze_panes = "A2"
    # 列宽
    for col_idx, col_name in enumerate(all_cols, start=1):
        max_len = len(col_name)
        for r in rows_data:
            max_len = max(max_len, len(str(r.get(col_name, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(max_len + 4, 56))

    # 应用样式
    _style_sheet(ws, len(all_cols), n_data, styles)

    # 合计行：浅灰底 + 加粗 + 上下边框 + 数字右对齐
    total_row_idx = n_data + 2
    ws.row_dimensions[total_row_idx].height = 26
    for col_idx, col_name in enumerate(all_cols, start=1):
        c = ws.cell(row=total_row_idx, column=col_idx)
        c.fill      = styles["total_fill"]
        c.font      = styles["total_font"]
        c.border    = styles["total_border"]
        c.alignment = _cell_align(c.value, styles)

    # "总Total Tokens用量" 列加粗，突出关键指标
    total_tokens_col = "总Total Tokens用量"
    if total_tokens_col in all_cols:
        token_col_idx = all_cols.index(total_tokens_col) + 1
        from openpyxl.styles import Font as _Font
        bold_token_font = _Font(bold=True, size=11, color="1A1A2E")
        for row in range(2, n_data + 2):
            c = ws.cell(row=row, column=token_col_idx)
            c.font = bold_token_font

    wb.save(out_path)
    log.info(f"汇总.xlsx 已写入 {out_path}（{len(rows_data)-1} 个账号）")
    return out_path


def _invoice_month_tag(month_value: str) -> str:
    """前端 month 传值（如 2026-04）规范化为 YYYY.MM。"""
    s = (month_value or "").strip()
    if not s:
        return ""
    m = re.search(r"(\d{4})\D+(\d{1,2})", s)
    if not m:
        return ""
    y = m.group(1)
    mo = int(m.group(2))
    return f"{y}.{mo:02d}"


def _billing_month_key(value: str) -> str:
    """把前端 month 或账单页 Date 文本规范化为 YYYY-MM。"""
    s = (value or "").strip()
    if not s:
        return ""

    m = re.search(r"(\d{4})\D+(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"

    month_names = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    m = re.search(r"\b([A-Za-z]+)\b\s+\d{1,2},\s*(\d{4})", s)
    if m:
        month = month_names.get(m.group(1).lower())
        if month:
            return f"{m.group(2)}-{month:02d}"

    return ""


def _billing_month_select_payload(value: str) -> dict:
    """生成账单页月份下拉可匹配的目标值和标签。"""
    key = _billing_month_key(value)
    if not key:
        return {}
    year_s, month_s = key.split("-", 1)
    year = int(year_s)
    month = int(month_s)
    short_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    full_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    labels = [
        f"{year}年{month}月",
        f"{year}年{month:02d}月",
        key,
        f"{short_names[month - 1]} {year}",
        f"{full_names[month - 1]} {year}",
    ]
    return {"value": key, "year": year, "month": month, "labels": labels}


def _month_distance_descending(current_value: str, target_value: str) -> Optional[int]:
    """按月份下拉降序列表计算从当前月到目标月需要 ArrowDown 的次数。"""
    current = _billing_month_key(current_value)
    target = _billing_month_key(target_value)
    if not current or not target:
        return None
    current_y, current_m = current.split("-", 1)
    target_y, target_m = target.split("-", 1)
    current_index = int(current_y) * 12 + int(current_m)
    target_index = int(target_y) * 12 + int(target_m)
    return current_index - target_index


def _normalize_status_text(s: str) -> str:
    """标准化状态文本，用于命名。"""
    v = (s or "").strip().lower()
    if not v:
        return ""
    # 中文状态映射到统一英文，避免文件名不可读
    if "退款" in v:
        return "refunded"
    if "支付" in v and ("未" in v or "待" in v):
        return "open"
    if "支付" in v:
        return "paid"
    if "草稿" in v:
        return "draft"
    if "作废" in v:
        return "void"
    if "无法收款" in v:
        return "uncollectible"
    # 英文常见状态
    if any(k in v for k in ("unpaid", "not paid", "past due", "payment due")):
        return "open"
    for k in ("open", "refunded", "void", "uncollectible", "draft", "paid"):
        if k in v:
            return k
    return _safe_filename(v)


def _normalize_invoice_url(url: str) -> str:
    """标准化发票 URL，避免同一链接因大小写/尾斜杠差异匹配失败。"""
    s = (url or "").strip()
    if not s:
        return ""
    try:
        p = urlsplit(s)
        return urlunsplit((
            p.scheme.lower(),
            p.netloc.lower(),
            p.path.rstrip("/"),
            p.query,
            "",
        ))
    except Exception:
        return s


_PAID_BILLING_STATUSES = frozenset({"paid", "refunded"})
"""可下载的"已支付"账单状态集合。

Stripe 的 `refunded` 表示账单先被支付后又退款，账单 PDF 真实存在且属于
已收款记录的一部分；与 `open`/`unpaid`/`void`/`draft`/`uncollectible`
等真正未付款状态不同。下载时一并保留，文件名沿用各自原始状态以示区分。
"""


def _filter_paid_billing_items(
    items: list[tuple],
    *,
    invoice_month: str = "",
) -> list[tuple[str, str]]:
    """只保留已支付（含已退款）账单，避免下载 Stripe 未支付账单/付款页。"""
    paid_items: list[tuple[str, str]] = []
    seen_urls: set[str] = set()
    selected_month = _billing_month_key(invoice_month)
    for item in items:
        if len(item) < 2:
            continue
        url, status = item[0], item[1]
        row_date = item[2] if len(item) >= 3 else ""
        normalized_status = _normalize_status_text(status)
        if normalized_status not in _PAID_BILLING_STATUSES:
            continue
        if selected_month and _billing_month_key(str(row_date)) != selected_month:
            continue
        normalized_url = _normalize_invoice_url(url)
        if not normalized_url or normalized_url in seen_urls:
            continue
        seen_urls.add(normalized_url)
        paid_items.append((url.strip(), normalized_status))
    return paid_items


def _unique_pdf_path(out_dir: Path, base_name: str) -> Path:
    """在 out_dir 下为 base_name（如 `xx-2026.04-paid.pdf`）生成不冲突的路径：
    首个不变；后续追加 `-2`、`-3`…… 直到不存在为止。
    """
    p = out_dir / base_name
    if not p.exists():
        return p
    stem = p.stem
    suffix = p.suffix
    i = 2
    while True:
        cand = out_dir / f"{stem}-{i}{suffix}"
        if not cand.exists():
            return cand
        i += 1


def _is_receipt_download_text(text: str) -> bool:
    """Stripe 已支付页同时有账单和收据按钮，收据不是我们要的文件。"""
    value = (text or "").strip().lower()
    return any(word in value for word in ("收据", "receipt"))


def _stripe_invoice_download_selectors() -> list[str]:
    """Stripe 页面下载选择器：先点账单/发票，最后才用通用 PDF 兜底。"""
    invoice_texts = (
        "下载账单",
        "下载发票",
        "Download invoice",
        "Invoice",
        "Invoice PDF",
    )
    selectors: list[str] = []
    for text in invoice_texts:
        selectors.extend([
            f'button:has-text("{text}")',
            f'a:has-text("{text}")',
            f'[role="button"]:has-text("{text}")',
        ])
    selectors.extend([
        ':text("PDF")',
        'a[href*=".pdf"]',
        'a[href*="/pdf"]',
        '[data-testid*="pdf"]',
        'a[download]',
    ])
    return selectors


_STATUS_JS = """
() => {
  const out = [];
  const STATUS_KEYS = ['paid','open','refunded','void','uncollectible','draft',
    '已支付','待支付','未支付','退款','草稿','作废','无法收款'];
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  for (const tr of document.querySelectorAll('tr')) {
    const a = tr.querySelector('a[href*="invoice.stripe.com"]');
    if (!a) continue;
    const tds = [...tr.querySelectorAll('td')];
    if (!tds.length) continue;
    const date = norm(tds[0]?.innerText || tds[0]?.textContent);
    let status = '';
    for (const td of tds) {
      if (td.classList.contains('capitalize')) {
        const t = norm(td.innerText || td.textContent);
        if (t) { status = t; break; }
      }
    }
    if (!status) {
      for (const td of tds) {
        const t = norm(td.innerText || td.textContent).toLowerCase();
        if (!t || t.length > 40) continue;
        if (STATUS_KEYS.some(k => t.includes(k.toLowerCase()))) { status = t; break; }
      }
    }
    out.push({ url: a.href || '', status, date });
  }
  return out;
}
"""

_BILLING_MONTH_SELECT_JS = """
async (target) => {
  if (!target || !target.value || !Array.isArray(target.labels)) return false;
  const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  const labels = target.labels.map(norm).filter(Boolean);
  const lowerLabels = labels.map(s => s.toLowerCase());
  const monthPattern = /\\d{4}\\s*年\\s*\\d{1,2}\\s*月|\\d{4}[-/.]\\d{1,2}|\\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\\s+\\d{4}\\b/i;
  const matches = (text, value = '') => {
    const candidates = [norm(text), norm(value)].filter(Boolean);
    return candidates.some(candidate => {
      const lower = candidate.toLowerCase();
      return lowerLabels.some(label => lower === label || lower.includes(label));
    });
  };
  const fire = el => {
    el.dispatchEvent(new Event('input', { bubbles: true }));
    el.dispatchEvent(new Event('change', { bubbles: true }));
  };

  for (const select of document.querySelectorAll('select')) {
    const options = [...select.options];
    const option = options.find(o => matches(o.textContent, o.value));
    if (!option) continue;
    select.value = option.value;
    select.selectedIndex = options.indexOf(option);
    fire(select);
    return true;
  }

  const clickables = [...document.querySelectorAll('button,[role="button"],[role="combobox"],[aria-haspopup="listbox"],[aria-haspopup="menu"],div[tabindex],span[tabindex]')];
  const trigger = clickables.find(el => {
    const text = norm(el.innerText || el.textContent);
    const aria = norm(`${el.getAttribute('aria-label') || ''} ${el.getAttribute('title') || ''}`);
    if (matches(text) || matches(aria)) return true;
    if (text.length <= 36 && monthPattern.test(text)) return true;
    return /账单|发票|invoice|billing|month|月份/i.test(aria) && text.length <= 50;
  });
  if (trigger && matches(trigger.innerText || trigger.textContent, trigger.getAttribute('aria-label') || trigger.getAttribute('title') || '')) {
    return true;
  }
  if (trigger) {
    const isOpen = trigger.getAttribute('aria-expanded') === 'true' || trigger.getAttribute('data-state') === 'open';
    if (!isOpen) {
      trigger.click();
      await sleep(450);
    }
  }

  const optionSelectors = '[role="option"],[role="menuitem"],[data-radix-collection-item],button,li,div[tabindex],span[tabindex]';
  for (let i = 0; i < 8; i += 1) {
    const options = [...document.querySelectorAll(optionSelectors)]
      .filter(el => el.offsetParent !== null || el.getClientRects().length > 0);
    const option = options.find(el => matches(el.innerText || el.textContent, el.getAttribute('aria-label') || el.getAttribute('title') || ''));
    if (option) {
      option.click();
      await sleep(700);
      return true;
    }
    await sleep(250);
  }
  return false;
}
"""

_BILLING_MONTH_PROBE_SELECT_JS = """
async (target) => {
  // 这个脚本只负责「发现」：打开下拉、读取选项列表、找到匹配文本。
  // 不再用 dispatchEvent 合成点击选项——合成事件不经过 React fiber，
  // 无法触发 React state update，需由 Playwright 真实点击来完成。
  const result = { selected: false, triggerText: '', matchedLabel: '', optionTexts: [], triggerCount: 0, reason: '' };
  if (!target || !target.value || !Array.isArray(target.labels)) {
    result.reason = 'invalid target';
    return result;
  }
  const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  const labels = target.labels.map(norm).filter(Boolean);
  const lowerLabels = labels.map(s => s.toLowerCase());
  const monthPattern = /\\d{4}\\s*年\\s*\\d{1,2}\\s*月|\\d{4}[-/.]\\d{1,2}|\\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\\s+\\d{4}\\b/i;
  const matches = (text, value = '') => {
    const candidates = [norm(text), norm(value)].filter(Boolean);
    return candidates.some(candidate => {
      const lower = candidate.toLowerCase();
      return lowerLabels.some(label => lower === label || lower.includes(label));
    });
  };
  const isVisible = el => !!(el && (el.offsetParent !== null || el.getClientRects().length > 0));
  const textOf = el => norm(`${el?.innerText || el?.textContent || ''} ${el?.getAttribute?.('aria-label') || ''} ${el?.getAttribute?.('title') || ''}`);
  const openTrigger = el => {
    // 仅用于打开下拉，不点选项
    if (!el) return;
    el.scrollIntoView?.({ block: 'center', inline: 'center' });
    el.focus?.();
    for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
      const event = type.startsWith('pointer')
        ? new PointerEvent(type, { bubbles: true, cancelable: true, pointerType: 'mouse', button: 0 })
        : new MouseEvent(type, { bubbles: true, cancelable: true, button: 0 });
      el.dispatchEvent(event);
    }
  };
  const optionSelector = '[role="option"],[role="menuitem"],[data-radix-collection-item],button,li,div[tabindex],span[tabindex]';
  const readOptions = () => [...document.querySelectorAll(optionSelector)]
    .filter(isVisible)
    .map(el => ({ el, text: textOf(el) }))
    .filter(item => item.text && item.text.length <= 80);
  const triggerSelector = 'button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"],button,[role="button"]';
  const triggers = [...document.querySelectorAll(triggerSelector)]
    .filter(isVisible)
    .map(el => ({ el, text: textOf(el), open: el.getAttribute('aria-expanded') === 'true' || el.getAttribute('data-state') === 'open' }))
    .filter(item => monthPattern.test(item.text) || item.open || /账单|发票|invoice|billing|month|月份/i.test(item.text));
  result.triggerCount = triggers.length;

  for (const item of triggers) {
    result.triggerText = item.text;
    if (matches(item.text)) {
      result.selected = true;
      result.matchedLabel = item.text;
      result.reason = 'trigger already selected';
      return result;
    }
    if (!item.open) {
      openTrigger(item.el);
      await sleep(500);
    }
    for (let i = 0; i < 8; i += 1) {
      const options = readOptions();
      result.optionTexts = options.map(o => o.text).slice(0, 20);
      const option = options.find(o => matches(o.text));
      if (option) {
        // 发现了匹配选项，记录文字供 Playwright 真实点击，不在 JS 侧点击
        result.matchedLabel = option.text;
        result.reason = 'matched portal option';
        return result;
      }
      await sleep(250);
    }
  }
  result.reason = result.optionTexts.length ? 'target option not found' : 'no dropdown options found';
  return result;
}
"""

_BILLING_MONTH_REFRESH_STATE_JS = """
(target) => {
  const state = {
    ready: false,
    selectedIndicator: false,
    triggerText: '',
    rowDates: [],
    targetRowDates: [],
    staleRowDates: [],
  };
  if (!target || !target.value || !Array.isArray(target.labels)) return state;
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  const labels = target.labels.map(norm).filter(Boolean).map(s => s.toLowerCase());
  const matchesTarget = text => {
    const lower = norm(text).toLowerCase();
    return labels.some(label => lower === label || lower.includes(label));
  };
  const monthKey = value => {
    const s = norm(value);
    const m = s.match(/(\\d{4})\\D+(\\d{1,2})/);
    if (!m) return '';
    return `${m[1]}-${String(Number(m[2])).padStart(2, '0')}`;
  };
  const triggerSelector = 'button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"],button,[role="button"]';
  const triggers = [...document.querySelectorAll(triggerSelector)];
  const trigger = triggers.find(el => matchesTarget(el.innerText || el.textContent || el.getAttribute('aria-label') || el.getAttribute('title') || ''));
  if (trigger) {
    state.selectedIndicator = true;
    state.triggerText = norm(trigger.innerText || trigger.textContent || trigger.getAttribute('aria-label') || trigger.getAttribute('title') || '');
  }
  const rows = [];
  for (const tr of document.querySelectorAll('tr')) {
    const a = tr.querySelector('a[href*="invoice.stripe.com"]');
    if (!a) continue;
    const firstCell = tr.querySelector('td');
    const date = norm(firstCell?.innerText || firstCell?.textContent || '');
    if (date) rows.push(date);
  }
  state.rowDates = rows;
  state.targetRowDates = rows.filter(date => monthKey(date) === target.value);
  state.staleRowDates = rows.filter(date => monthKey(date) && monthKey(date) !== target.value);
  state.ready = (state.selectedIndicator || state.targetRowDates.length > 0) && state.staleRowDates.length === 0;
  return state;
}
"""

_BILLING_MONTH_CURRENT_JS = """
(target) => {
  if (!target || !Array.isArray(target.labels)) return false;
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim().toLowerCase();
  const labels = target.labels.map(norm).filter(Boolean);
  const candidates = [...document.querySelectorAll('button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"]')];
  return candidates.some(el => {
    const text = norm(el.innerText || el.textContent || '');
    const aria = norm(`${el.getAttribute('aria-label') || ''} ${el.getAttribute('title') || ''}`);
    return labels.some(label => text === label || text.includes(label) || aria.includes(label));
  });
}
"""

_BILLING_MONTH_TRIGGER_TEXT_JS = """
() => {
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  const monthPattern = /\\d{4}\\s*年\\s*\\d{1,2}\\s*月|\\d{4}[-/.]\\d{1,2}|\\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\\s+\\d{4}\\b/i;
  const candidates = [...document.querySelectorAll('button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"]')];
  const trigger = candidates.find(el => monthPattern.test(norm(el.innerText || el.textContent || '')));
  return trigger ? norm(trigger.innerText || trigger.textContent || '') : '';
}
"""

_BILLING_MONTH_TRIGGER_SELECTOR = 'button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"]'
_BILLING_MONTH_OPTION_SELECTOR = '[role="option"],[role="menuitem"],[data-radix-collection-item],button'

_BILLING_URLS = [
    "https://cursor.com/dashboard/billing",   # 月份下拉控件可靠，优先尝试
    "https://cursor.com/settings/billing",    # 备用
]


_INVOICE_TRIGGER_EXACT_MONTH_RE = re.compile(
    r"^\s*\d{4}\s*年\s*\d{1,2}\s*月\s*$|"
    r"^\s*\d{4}[-/.]\d{1,2}\s*$|"
    r"^\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s+\d{4}\s*$",
    re.I,
)
_CYCLE_OR_NOISE_RE = re.compile(
    r"Cycle Starting|Adjust plan|Manage in Stripe|Cancel|Upgrade Now|Cursor navigation|User menu",
    re.I,
)

_BILLING_MONTH_OPTIONS_VISIBLE_JS = """
(target) => {
  // 探测目标月份选项当前是否已经在 DOM 中可见。
  // probe 阶段用合成 PointerEvent 可能已经打开了下拉；若我们再 click trigger
  // 会把它"toggle 关闭"，所以点击 trigger 之前先用此脚本判断。
  if (!target || !Array.isArray(target.labels)) return false;
  const norm = s => (s || '').replace(/\\s+/g, ' ').trim();
  const lowerLabels = target.labels.map(s => norm(s).toLowerCase()).filter(Boolean);
  if (!lowerLabels.length) return false;
  const isVisible = el => !!(el && (el.offsetParent !== null || el.getClientRects().length > 0));
  const optionSelector = '[role="option"],[role="menuitem"],[data-radix-collection-item]';
  for (const el of document.querySelectorAll(optionSelector)) {
    if (!isVisible(el)) continue;
    const text = norm(el.innerText || el.textContent).toLowerCase();
    if (!text || text.length > 80) continue;
    if (/cycle starting|cancel|adjust plan|manage in stripe/i.test(text)) continue;
    if (lowerLabels.some(label => text === label)) return true;
  }
  return false;
}
"""

_BILLING_PAGE_READY_JS = """
() => {
  // 严格的账单页就绪判定（避免被导航栏 button/a 提前误判）：
  // 1) 已出现 Stripe 发票链接（当前筛选下有账单）
  // 2) 已出现精确月份 trigger（可进行月份切换）
  // 3) 已出现明确空态文案（无账单）
  if (document.querySelector('a[href*="invoice.stripe.com"]')) return true;
  const monthRe = /^\\s*\\d{4}\\s*年\\s*\\d{1,2}\\s*月\\s*$|^\\s*\\d{4}[-/.]\\d{1,2}\\s*$|^\\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\\s+\\d{4}\\s*$/i;
  const triggerSelector = 'button[aria-expanded][aria-controls],button[aria-haspopup],[role="combobox"],button';
  for (const el of document.querySelectorAll(triggerSelector)) {
    const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
    if (text && monthRe.test(text)) return true;
  }
  const pageText = (document.body?.innerText || '').replace(/\\s+/g, ' ');
  if (/no invoices|no past invoices|没有账单|没有发票|暂无账单|暂无发票/i.test(pageText)) {
    return true;
  }
  return false;
}
"""


async def _wait_billing_page_ready(page, *, timeout_ms: int = 25000) -> bool:
    """等待账单区域真正渲染完成。

    与通用 ``button, a[href]`` 不同，此函数只在账单核心元素出现时返回 True，
    防止页面尚未 hydrate 就进入月份切换流程。
    """
    try:
        await page.wait_for_function(_BILLING_PAGE_READY_JS, timeout=timeout_ms)
        await page.wait_for_timeout(600)
        return True
    except Exception:
        return False


async def _select_billing_month_via_playwright(page, payload: dict) -> bool:
    """用 Playwright 严格定位 Invoices 表头的月份过滤器并点击切换。

    页面上存在多个看起来像"月份下拉"的控件（例如订阅周期管理弹窗里的
    "Cycle Starting 2026年X月XX日" 选项），但它们和 Invoices 列表完全无关。
    真正的过滤器特征是：trigger 文本是简短且精确的 "YYYY年M月" 格式，
    选项文本同样是简短的月份。所以这里采用：
      1) 用 ^YYYY年M月$ 的精确正则只匹配简短月份按钮
      2) 排除任何包含 "Cycle Starting / Adjust plan / Cancel" 等噪声词的元素
      3) 由 Playwright 真实点击触发 React 事件
    """
    labels = [str(v) for v in payload.get("labels", []) if v]
    if not labels:
        return False
    target_value = str(payload.get("value", ""))

    try:
        probe = await page.evaluate(_BILLING_MONTH_PROBE_SELECT_JS, payload)
        if isinstance(probe, dict):
            options = probe.get("optionTexts") or []
            if options:
                log.info(
                    "账单页月份下拉探测(诊断): "
                    f"trigger={str(probe.get('triggerText', ''))[:40]!r}, "
                    f"options={options[:12]}"
                )
    except Exception:
        pass

    # probe 阶段可能用合成 PointerEvent 已经打开了 Radix portal 下拉；
    # 若再 click trigger 会切换关闭，导致下面找不到选项。所以先探测：
    options_already_visible = False
    try:
        options_already_visible = bool(
            await page.evaluate(_BILLING_MONTH_OPTIONS_VISIBLE_JS, payload)
        )
    except Exception:
        options_already_visible = False

    try:
        all_triggers = page.locator("button").filter(
            has_text=_INVOICE_TRIGGER_EXACT_MONTH_RE
        )
        trigger_count = await all_triggers.count()
        if trigger_count == 0 and not options_already_visible:
            log.info("账单页未找到精确 'YYYY年M月' 格式的过滤器按钮")
            return False

        trigger = None
        for i in range(min(trigger_count, 5)):
            candidate = all_triggers.nth(i)
            try:
                text = (await candidate.inner_text(timeout=1500)).strip()
            except Exception:
                continue
            if _CYCLE_OR_NOISE_RE.search(text):
                continue
            if not _INVOICE_TRIGGER_EXACT_MONTH_RE.match(text):
                continue
            trigger = candidate
            log.info(f"账单页过滤器定位成功: text={text!r}")
            break

        if trigger is None and not options_already_visible:
            log.info("账单页 'YYYY年M月' 候选按钮均不通过过滤")
            return False

        if trigger is not None and not options_already_visible:
            try:
                current_text = (await trigger.inner_text(timeout=1500)).strip()
            except Exception:
                current_text = ""
            if current_text and any(
                label.lower() in current_text.lower() for label in labels
            ):
                log.info(f"账单页过滤器已是目标月份: current={current_text!r}")
                return True

        if options_already_visible:
            log.info("账单页下拉已打开（probe 已展开），跳过 trigger click 直接选项点击")
        else:
            await trigger.scroll_into_view_if_needed(timeout=2000)
            await trigger.click(timeout=3000)
            await page.wait_for_timeout(700)
    except Exception as e:
        log.info(f"账单页过滤器点击触发器失败: {e}")
        return False

    option_selector = (
        '[role="option"],[role="menuitem"],[data-radix-collection-item]'
    )
    for label in labels:
        exact_re = re.compile(rf"^\s*{re.escape(label)}\s*$")
        try:
            option = (
                page.locator(option_selector)
                .filter(has_text=exact_re)
                .filter(has_not_text=_CYCLE_OR_NOISE_RE)
                .first
            )
            if await option.count() == 0:
                continue
            await option.scroll_into_view_if_needed(timeout=1500)
            await option.click(timeout=3000)
            await page.wait_for_timeout(800)
            log.info(f"账单页过滤器已点击月份选项: {label!r}")
            return True
        except Exception:
            continue

    for label in labels:
        try:
            option = (
                page.locator(option_selector)
                .filter(has_text=label)
                .filter(has_not_text=_CYCLE_OR_NOISE_RE)
                .first
            )
            if await option.count() == 0:
                continue
            await option.scroll_into_view_if_needed(timeout=1500)
            await option.click(timeout=3000)
            await page.wait_for_timeout(800)
            log.info(f"账单页过滤器已点击月份选项(模糊): {label!r}")
            return True
        except Exception:
            continue

    try:
        current_text = str(await page.evaluate(_BILLING_MONTH_TRIGGER_TEXT_JS) or "")
        distance = _month_distance_descending(current_text, target_value)
        if distance is None or distance < 0 or distance > 36:
            return False
        for _ in range(distance):
            await page.keyboard.press("ArrowDown")
            await page.wait_for_timeout(80)
        await page.keyboard.press("Enter")
        await page.wait_for_timeout(800)
        return True
    except Exception:
        return False


async def _wait_for_billing_month_refresh(page, payload: dict, *, timeout_ms: int = 20000) -> bool:
    """点击月份后等待账单列表刷新，避免马上读到旧月份的发票行。"""
    attempts = max(1, timeout_ms // 500)
    last_state: dict = {}
    for _ in range(attempts):
        try:
            state = await page.evaluate(_BILLING_MONTH_REFRESH_STATE_JS, payload)
            if isinstance(state, dict):
                last_state = state
                if state.get("ready"):
                    row_dates = state.get("rowDates") or []
                    log.info(
                        "账单页月份刷新确认: "
                        f"trigger={str(state.get('triggerText', ''))[:40]!r}, "
                        f"rows={row_dates[:8]}"
                    )
                    return True
        except Exception:
            pass
        await page.wait_for_timeout(500)
    log.warning(
        "账单页月份切换后列表未确认刷新: "
        f"target={payload.get('value')}, "
        f"trigger={str(last_state.get('triggerText', ''))[:40]!r}, "
        f"rows={(last_state.get('rowDates') or [])[:8]}, "
        f"stale={(last_state.get('staleRowDates') or [])[:8]}"
    )
    return False


async def _select_billing_month_in_ctx(page, invoice_month: str) -> bool:
    """先把账单页月份控件切到用户选择的月份，再解析列表。"""
    payload = _billing_month_select_payload(invoice_month)
    if not payload:
        return False
    try:
        selected = await _select_billing_month_via_playwright(page, payload)
        if not selected:
            selected = bool(await page.evaluate(_BILLING_MONTH_SELECT_JS, payload))
        if selected:
            refreshed = await _wait_for_billing_month_refresh(page, payload)
            if not refreshed:
                return False
            log.info(f"账单页月份已切换到 {payload['value']}")
        else:
            log.warning(f"账单页未找到可切换到 {payload['value']} 的月份控件，继续使用列表日期过滤兜底")
        return selected
    except Exception as e:
        log.warning(f"账单页月份切换失败: {e}")
        return False


async def _fetch_billing_items_in_ctx(page, invoice_month: str = "") -> list[tuple[str, str, str]]:
    """在已有 page 对象上抓取账单页状态列表，返回 [(url, status, date), ...]。"""
    items: list[tuple[str, str, str]] = []
    requested_month = _billing_month_key(invoice_month)
    for billing_url in _BILLING_URLS:
        try:
            await page.goto(billing_url, wait_until="load", timeout=20000)
            ready = await _wait_billing_page_ready(page)
            if not ready:
                log.info(f"账单页核心区域未就绪，跳过 URL: {billing_url}")
                continue
            if requested_month:
                selected = await _select_billing_month_in_ctx(page, invoice_month)
                if not selected:
                    continue
        except Exception:
            continue
        found_rows: list[dict] = await page.evaluate(_STATUS_JS)
        items = [
            (
                r["url"],
                _normalize_status_text(str(r.get("status", ""))),
                str(r.get("date", "")),
            )
            for r in (found_rows or [])
            if isinstance(r, dict)
            and str(r.get("url", "")).startswith("http")
            and _normalize_status_text(str(r.get("status", "")))
        ]
        log.info(f"账单页 {billing_url}: {len(found_rows or [])} 行, 可用 {len(items)}")
        if items:
            for u, s, d in items:
                log.info(f"  date={d!r} status={s!r} url={u[:70]}")
            break
    return items


async def _download_account_all_pdfs(
    cookie_val: str,
    out_dir: Path,
    email_tag: str,
    month_tag: str,
    email: str,
    invoice_month: str = "",
    sem: Optional[asyncio.Semaphore] = None,
    browser=None,
) -> dict[str, str]:
    """单账号抓状态 + 下载所有 PDF。

    默认复用外部传入的 browser（单 Chromium 多 Context 并发模型）。
    若 browser 为空，退化为兼容模式：函数内部自建 browser。
    """

    async def _run_with_browser(active_browser) -> dict[str, str]:
        out_dir.mkdir(parents=True, exist_ok=True)
        pdf_files: dict[str, str] = {}
        try:
            ctx = await active_browser.new_context(accept_downloads=True)
            try:
                await ctx.add_cookies([{
                    "name": "WorkosCursorSessionToken",
                    "value": cookie_val,
                    "domain": "cursor.com",
                    "path": "/",
                    "httpOnly": True,
                    "secure": True,
                }])

                # ── Step 1: 抓账单状态 ──────────────────────────────
                status_page = await ctx.new_page()
                try:
                    items = await _fetch_billing_items_in_ctx(status_page, invoice_month=invoice_month or month_tag)
                finally:
                    await status_page.close()

                paid_items = _filter_paid_billing_items(
                    items,
                    invoice_month=invoice_month or month_tag,
                )
                if not paid_items:
                    log.warning(
                        f"[{email}] 账单页未找到 {month_tag} 的已支付发票行，跳过"
                    )
                    return {}

                log.info(
                    f"[{email}] 账单页状态抓取成功: 总计 {len(items)} 条，"
                    f"{month_tag} 已支付 {len(paid_items)} 条"
                )

                # ── Step 2: 下载每张 PDF（复用同一 ctx，无需重新认证）──
                for idx, (hosted_url, status_raw) in enumerate(paid_items, start=1):
                    status_tag = _safe_filename(status_raw)
                    if not status_tag:
                        log.warning(f"[{email}] 账单 {idx} 无 status，跳过")
                        continue

                    base_name = _safe_filename(f"{email_tag}-{month_tag}-{status_tag}.pdf")
                    save_path = _unique_pdf_path(out_dir, base_name)
                    fname = save_path.name

                    log.info(f"[{email}] 下载已支付账单 {idx}/{len(paid_items)}: {fname}")
                    ok = False
                    for _attempt in range(1, 4):
                        ok = await _stripe_page_click_pdf(hosted_url, save_path, ctx)
                        if ok:
                            break
                        log.warning(f"[{email}] 账单 {idx} 第 {_attempt} 次失败，重试...")
                        await asyncio.sleep(3)

                    if ok:
                        pdf_files[str(idx)] = fname
                        log.info(f"[{email}] 账单 PDF 下载成功: {fname}")
                    else:
                        log.warning(f"[{email}] 账单 {idx} 重试 3 次均失败，跳过")
            finally:
                await ctx.close()
        except Exception as e:
            log.warning(f"[{email}] 浏览器账单下载整体异常: {e}")

        return pdf_files

    async def _run() -> dict[str, str]:
        if browser is not None:
            return await _run_with_browser(browser)
        from patchright.async_api import async_playwright

        async with async_playwright() as pw:
            local_browser = await pw.chromium.launch(headless=True)
            try:
                return await _run_with_browser(local_browser)
            finally:
                await local_browser.close()

    if sem is not None:
        async with sem:
            return await _run()
    return await _run()


async def _billing_page_get_stripe_items(cookie_val: str, invoice_month: str = "") -> list[tuple[str, str, str]]:
    """用 patchright + Cookie 访问 Cursor 账单页，提取发票链接及状态。

    返回: [(hosted_invoice_url, status, date), ...]
    status 为账单列表状态列文本（会转小写）。
    """
    from patchright.async_api import async_playwright

    items: list[tuple[str, str, str]] = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context()
        await ctx.add_cookies([{
            "name": "WorkosCursorSessionToken",
            "value": cookie_val,
            "domain": "cursor.com",
            "path": "/",
            "httpOnly": True,
            "secure": True,
        }])
        page = await ctx.new_page()
        try:
            # 尝试几个可能的账单页 URL，每个 URL 只尝试一次
            for billing_url in _BILLING_URLS:
                try:
                    await page.goto(billing_url, wait_until="load", timeout=20000)
                    ready = await _wait_billing_page_ready(page)
                    if not ready:
                        log.info(f"账单页核心区域未就绪，跳过 URL: {billing_url}")
                        continue
                except Exception:
                    continue
                if _billing_month_key(invoice_month):
                    selected = await _select_billing_month_in_ctx(page, invoice_month)
                    if not selected:
                        continue

                # 同一行内抓 URL / Date / Status
                found_rows: list[dict[str, str]] = await page.evaluate(_STATUS_JS)
                items = [
                    (
                        row.get("url", ""),
                        _normalize_status_text(str(row.get("status", ""))),
                        str(row.get("date", "")),
                    )
                    for row in (found_rows or [])
                    if isinstance(row, dict)
                    and str(row.get("url", "")).startswith("http")
                    and _normalize_status_text(str(row.get("status", "")))
                ]
                log.info(
                    f"账单页 {billing_url}: 行数={len(found_rows or [])}, "
                    f"可用={len(items)}"
                )
                if items:
                    for u, s, d in items:
                        log.info(f"  行 date={d!r} status={s!r} url={u[:70]}")
                    break
        except Exception as e:
            log.warning(f"账单页提取链接失败: {e}")
        finally:
            await browser.close()

    return items


async def _stripe_page_click_pdf(
    hosted_url: str,
    save_path: Path,
    ctx,  # playwright BrowserContext
) -> bool:
    """在 Stripe 发票页面点击 PDF 按钮下载。

    Stripe 是 React SPA，需要等 networkidle 后再等额外 4s。
    PDF 按钮可能是 <a>/<button>/<div> 任意元素，用宽泛 :text 匹配。
    """
    page = await ctx.new_page()
    try:
        # Stripe 是 React SPA，用 "load" 而非 "networkidle"
        # （networkidle 会因后台持续请求而超时）
        await page.goto(hosted_url, wait_until="load", timeout=30000)
        # 等 React 渲染完：先等 button 出现（最多 8s），再额外等 1s
        try:
            await page.wait_for_selector("button, a[href]", timeout=8000)
        except Exception:
            pass
        await page.wait_for_timeout(1500)

        import requests as _req

        # ── 诊断1：页面标题和当前 URL ──
        title = await page.title()
        log.info(f"Stripe 页面加载完成: title={title!r}, url={page.url[:80]}")

        # ── 诊断2：所有含 PDF 文字的元素（不限 tag）──
        pdf_els = await page.evaluate("""
            () => [...document.querySelectorAll('*')]
                .filter(e => e.children.length === 0 && e.innerText &&
                             e.innerText.toUpperCase().includes('PDF'))
                .slice(0, 10)
                .map(e => ({
                    tag:  e.tagName,
                    text: e.innerText.trim().slice(0, 80),
                    href: e.href  || '',
                    cls:  e.className.slice(0, 60),
                    id:   e.id    || '',
                }))
        """)
        log.info(f"Stripe 含PDF元素({len(pdf_els)}个): {pdf_els}")

        # ── 诊断3：所有可点击元素（a/button）前10个 ──
        clickable = await page.evaluate("""
            () => [...document.querySelectorAll('a, button')]
                .slice(0, 10)
                .map(e => ({
                    tag:  e.tagName,
                    text: e.innerText.trim().slice(0, 50),
                    href: e.href || '',
                }))
        """)
        log.info(f"Stripe 可点击元素(前10): {clickable}")

        # ── 选择器按优先级依次尝试 ──
        selectors = _stripe_invoice_download_selectors()

        for sel in selectors:
            try:
                el = page.locator(sel).first
                count = await el.count()
                log.info(f"  selector {sel!r} → {count} 个")
                if count == 0:
                    continue

                href = await el.get_attribute("href") or ""
                tag  = await el.evaluate("e => e.tagName")
                text = await el.inner_text()
                log.info(f"  命中: tag={tag}, text={text[:50]!r}, href={href[:80]!r}")
                if _is_receipt_download_text(text):
                    log.info(f"  跳过收据下载按钮: text={text[:50]!r}")
                    continue

                if href.startswith("http") and (".pdf" in href.lower() or "/pdf" in href.lower()):
                    resp = _req.get(href, timeout=30, stream=True,
                                    headers={"User-Agent": "Mozilla/5.0"})
                    resp.raise_for_status()
                    save_path.parent.mkdir(parents=True, exist_ok=True)
                    with save_path.open("wb") as f:
                        for chunk in resp.iter_content(8192):
                            if chunk:
                                f.write(chunk)
                    if save_path.exists() and save_path.read_bytes()[:4] == b"%PDF":
                        log.info(f"PDF href 下载成功: {save_path.name}")
                        return True

                log.info(f"  点击元素等待 download 事件…")
                try:
                    async with page.expect_download(timeout=15000) as dl_info:
                        await el.click()
                    download = await dl_info.value
                    save_path.parent.mkdir(parents=True, exist_ok=True)
                    await download.save_as(str(save_path))
                    if save_path.exists() and save_path.read_bytes()[:4] == b"%PDF":
                        log.info(f"PDF 点击下载成功: {save_path.name}")
                        return True
                except Exception as click_err:
                    log.info(f"  download 事件未触发({click_err})，检查新 tab…")
                    await page.wait_for_timeout(2000)
                    new_pages = [p for p in ctx.pages if p != page and ".pdf" in p.url.lower()]
                    if new_pages:
                        pdf_url = new_pages[0].url
                        log.info(f"  新 tab PDF URL: {pdf_url[:80]}")
                        await new_pages[0].close()
                        resp = _req.get(pdf_url, timeout=30, stream=True,
                                        headers={"User-Agent": "Mozilla/5.0"})
                        resp.raise_for_status()
                        save_path.parent.mkdir(parents=True, exist_ok=True)
                        with save_path.open("wb") as f:
                            for chunk in resp.iter_content(8192):
                                if chunk:
                                    f.write(chunk)
                        if save_path.exists() and save_path.read_bytes()[:4] == b"%PDF":
                            return True

            except Exception as e:
                log.info(f"  selector {sel!r} 异常: {e}")
                continue

        log.warning(
            f"Stripe 页面未找到PDF按钮。"
            f"title={title!r}, PDF元素={len(pdf_els)}个, 可点击={len(clickable)}个"
        )
        return False
    except Exception as e:
        log.warning(f"Stripe 页面访问失败: {e}")
        return False
    finally:
        await page.close()


def _download_invoices_all(
    snapshots: list,
    accounts: list,
    out_root: Path,
    *,
    manager,
    invoice_month: str,
    concurrency: int,
    progress_cb=None,
) -> dict:
    """并发下载所有账号账单（单 Chromium + 多 Context）。

    返回 {email: {idx_str: filename}} 映射。
    """
    from .api_client import _split_session_token

    month_tag = _invoice_month_tag(invoice_month)
    if not month_tag:
        return {}

    acc_by_email = {a.email: a for a in accounts}
    def _cb(email: str, phase: str, msg: str = "") -> None:
        if progress_cb:
            try:
                progress_cb(email, phase, msg)
            except Exception:
                pass

    # 页面加载阶段对资源敏感。INVOICE_ACTIVE_CONTEXT_LIMIT 用于限制“同时活跃
    # 的 billing context 数”；当该值 <= 0 时，表示不额外限制（仅受并发配置约束）。
    from .config import SETTINGS as _SETTINGS
    active_limit_cfg = int(getattr(_SETTINGS, "invoice_active_context_limit", 3))
    if active_limit_cfg > 0:
        max_parallel = max(1, min(concurrency, len(snapshots), active_limit_cfg))
    else:
        max_parallel = max(1, min(concurrency, len(snapshots)))
    log.info(
        f"[并发诊断] _download_invoices_all 启动: "
        f"snapshots={len(snapshots)}, concurrency={concurrency}, "
        f"active_context_limit={max_parallel}, "
        f"INVOICE_ACTIVE_CONTEXT_LIMIT={active_limit_cfg}"
    )

    async def _one_account(snap, browser, sem: asyncio.Semaphore) -> tuple[str, dict]:
        """复用同一个 Chromium，为单账号创建独立 context 下载账单。"""
        import time as _time
        _t0 = _time.time()
        acc = acc_by_email.get(snap.email)
        if acc is None:
            return snap.email, {}
        try:
            token = await asyncio.to_thread(manager.get_valid_token, acc)
        except Exception as e:
            log.warning(f"[{snap.email}] 获取 token 失败，跳过账单: {e}")
            return snap.email, {}

        cookie_val, _ = _split_session_token(token)
        if not cookie_val:
            log.warning(f"[{snap.email}] cookie 为空，跳过账单")
            return snap.email, {}

        email_tag = _safe_filename(snap.email)
        out_dir = out_root / email_tag / "invoices"

        _cb(snap.email, "invoice", "下载账单中...")
        try:
            pdf_files = await _download_account_all_pdfs(
                cookie_val, out_dir, email_tag, month_tag, snap.email,
                invoice_month=invoice_month,
                sem=sem,
                browser=browser,
            )
        except Exception as e:
            log.warning(f"[{snap.email}] 账单下载异常: {e}")
            pdf_files = {}

        inv_count = len(pdf_files)
        log.info(f"[{snap.email}] 账单阶段完成: {inv_count} 份, dt={_time.time()-_t0:.3f}s")
        _cb(snap.email, "done", f"账单 {inv_count} 份" if inv_count else "")
        return snap.email, pdf_files

    async def _run() -> dict:
        from patchright.async_api import async_playwright

        out: dict = {}
        sem = asyncio.Semaphore(max_parallel)
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            try:
                tasks = [
                    asyncio.create_task(_one_account(snap, browser, sem))
                    for snap in snapshots
                ]
                for task in asyncio.as_completed(tasks):
                    try:
                        email, pdf_files = await task
                        out[email] = pdf_files
                    except Exception as e:
                        log.warning(f"账单下载任务异常: {e}")
            finally:
                try:
                    await browser.close()
                except Exception as e:
                    log.warning(f"关闭共享浏览器异常: {e}")
        return out

    try:
        return asyncio.run(_run())
    except Exception as e:
        log.warning(f"账单并发下载整体失败: {e}")
        return {}


def _download_pdfs_via_billing_page(
    access_token: str,
    out_dir: Path,
    *,
    email_tag: str = "account",
    invoice_month: str = "",
    stripe_items: Optional[list[tuple]] = None,
) -> dict[str, str]:
    """完整浏览器流程：
      Cursor 账单页 (View 链接) → Stripe 发票页 → 点击 PDF 按钮 → 下载。
    返回 {索引: filename} 映射（filename 包含状态名）。
    """
    import asyncio
    from .api_client import _split_session_token

    cookie_val, _ = _split_session_token(access_token)
    if not cookie_val:
        return {}

    async def _run() -> dict[str, str]:
        from patchright.async_api import async_playwright

        pdf_files: dict[str, str] = {}
        month_tag = _invoice_month_tag(invoice_month)
        if not month_tag:
            log.warning("账单月份为空或格式非法，跳过浏览器账单下载")
            return {}

        # Step1: 拿到所有 Stripe 发票 URL + 状态
        items = stripe_items if stripe_items is not None else await _billing_page_get_stripe_items(cookie_val, invoice_month=invoice_month)
        paid_items = _filter_paid_billing_items(items, invoice_month=month_tag)
        if not paid_items:
            log.warning(f"账单页未找到 {month_tag} 的已支付发票链接")
            return {}

        out_dir.mkdir(parents=True, exist_ok=True)
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            ctx = await browser.new_context(accept_downloads=True)

            for idx, (hosted_url, status_raw) in enumerate(paid_items, start=1):
                status_tag = _safe_filename(_normalize_status_text(status_raw))
                if not status_tag:
                    log.warning(f"账单 {idx} 缺少 status，跳过")
                    continue
                base_name = _safe_filename(f"{email_tag}-{month_tag}-{status_tag}.pdf")
                save_path = _unique_pdf_path(out_dir, base_name)
                fname = save_path.name

                log.info(
                    f"浏览器下载已支付账单 {idx}/{len(paid_items)}: "
                    f"month={month_tag}, status={status_tag}, url={hosted_url[:80]}"
                )
                ok = False
                for _attempt in range(1, 4):  # 最多重试 3 次
                    ok = await _stripe_page_click_pdf(hosted_url, save_path, ctx)
                    if ok:
                        break
                    log.warning(f"账单 {idx} 第 {_attempt} 次下载失败，等待后重试...")
                    await asyncio.sleep(3)
                if ok:
                    pdf_files[str(idx)] = fname
                    log.info(f"账单 PDF 下载成功: {fname}")
                else:
                    log.warning(f"账单 {idx} 重试 3 次均失败，跳过")

            await browser.close()

        return pdf_files

    try:
        return asyncio.run(_run())
    except Exception as e:
        log.warning(f"浏览器账单下载整体失败: {e}")
        return {}


def _render_pdf_images_to_sheet(ws, pdf_paths: list[Path], email: str) -> None:
    """把 PDF 发票第一页渲染成图片，嵌入 Excel sheet（无表头）。
    每个账号前插一行 email 标签，后接发票图片。
    需要 pymupdf（pip install pymupdf）。
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        import pymupdf as fitz   # 新版包名（pymupdf >= 1.24）
        has_fitz = True
    except ImportError:
        try:
            import fitz            # 旧版兼容（pip install pymupdf 仍可用 fitz）
            has_fitz = True
        except ImportError:
            has_fitz = False
            log.warning("pymupdf 未安装，账单 sheet 将仅显示文件名（pip install pymupdf）")

    valid_pdfs = [p for p in pdf_paths if p.exists() and p.stat().st_size > 512]
    if not valid_pdfs:
        return

    # 用当前 sheet 最后有内容的行之后插入
    cur_row = ws.max_row + (2 if ws.max_row > 1 else 1)

    # 账号标签行
    cell = ws.cell(row=cur_row, column=1, value=f"📧 {email}")
    cell.font = Font(bold=True, size=11, color="1155CC")
    cell.fill = PatternFill("solid", fgColor="EEF4FF")
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[cur_row].height = 20
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 100)
    cur_row += 1

    for pdf_path in valid_pdfs:
        if not has_fitz:
            # 无 pymupdf → 只写文件名
            ws.cell(row=cur_row, column=1, value=f"[PDF] {pdf_path.name}")
            cur_row += 1
            continue

        try:
            doc = fitz.open(str(pdf_path))
            page = doc[0]
            # 150 DPI：清晰度与文件大小的最佳平衡（官方推荐用 dpi= 参数）
            pix = page.get_pixmap(dpi=150, alpha=False)
            img_bytes = pix.tobytes("png")
            doc.close()

            from io import BytesIO
            img = XLImage(BytesIO(img_bytes))
            # 目标宽度 680px（Excel 列宽约 97 字符），等比缩放
            target_w = 680
            scale = target_w / pix.width
            img.width  = target_w
            img.height = int(pix.height * scale)
            img.anchor = f"A{cur_row}"
            ws.add_image(img)

            # 预留足够行高（Excel 行高单位是 pt，1px ≈ 0.75pt）
            rows_needed = max(int(img.height * 0.75 / 15) + 1, 2)
            for r in range(cur_row, cur_row + rows_needed):
                ws.row_dimensions[r].height = 15
            cur_row += rows_needed + 1  # 图片后空一行

        except Exception as e:
            log.warning(f"PDF 渲染失败 {pdf_path.name}: {e}")
            ws.cell(row=cur_row, column=1, value=f"[PDF 渲染失败] {pdf_path.name}: {e}")
            cur_row += 1


def export_per_account(
    accounts: list[Account],
    snapshots: list[AccountSnapshot],
    out_root: Path,
    *,
    manager: Optional[TokenManager] = None,
    with_raw: bool = True,
    with_invoices: bool = True,
    with_detail_xlsx: bool = True,
    with_full_summary_xlsx: bool = True,
    with_summary: bool = True,
    invoice_month: str = "",
    start_date: str = "",
    end_date: str = "",
    progress_cb=None,          # callable(email: str, phase: str, msg: str = "")
    browser_semaphore=None,    # threading.Semaphore，限制并发浏览器实例数
) -> dict[str, dict[str, Any]]:
    """每个账号一个子目录：

        {out_root}/
          {email}/
            {email}.xlsx       ← Excel（账号概览 / 使用明细，可选）
            raw.json           ← 原始 JSON（可选）
            invoices/
              {id}.pdf ...     ← 发票 PDF（可选）
          汇总.xlsx             ← Token 汇总（可选）
          _summary.xlsx        ← 全账号明细汇总（可选，默认开启）

    progress_cb(email, phase, msg) 在关键节点回调：
      invoice  → 开始下载账单
      excel    → 开始生成报表（with_detail_xlsx=True 时）
      done     → 该账号全部完成
      summary  → 开始生成汇总（email 为空字符串）
    """
    def _cb(email: str, phase: str, msg: str = "") -> None:
        if progress_cb:
            try:
                progress_cb(email, phase, msg)
            except Exception:
                pass

    import concurrent.futures as _cf
    import threading as _threading

    mgr = manager or get_default_manager()
    out_root = Path(out_root)
    out_root.mkdir(parents=True, exist_ok=True)

    # ── Phase 1：并发下载账单 PDF ──────────────────────────────────
    # 单 Chromium + 多 Context 模型：并发数代表“同时活跃 context 数”。
    # 该并发与登录并发解耦，直接读取配置值。
    from .config import SETTINGS as _SETTINGS
    concurrency = max(1, _SETTINGS.invoice_download_concurrency)

    all_pdf_files: dict[str, dict[str, str]] = {}
    if with_invoices and snapshots:
        all_pdf_files = _download_invoices_all(
            snapshots, accounts, out_root,
            manager=mgr,
            invoice_month=invoice_month,
            concurrency=concurrency,
            progress_cb=progress_cb,
        )

    # ── Phase 2：Excel / raw JSON（不涉及浏览器，可安全并发）────────
    summary: dict[str, dict[str, Any]] = {}
    _lock = _threading.Lock()

    def _process_one(snap: AccountSnapshot) -> None:
        acc_dir = out_root / _safe_filename(snap.email)
        acc_dir.mkdir(parents=True, exist_ok=True)

        pdf_files = all_pdf_files.get(snap.email, {})

        xlsx_path = None
        if with_detail_xlsx:
            _cb(snap.email, "excel", "生成报表...")
            xlsx_path = acc_dir / f"{_safe_filename(snap.email)}.xlsx"
            export_xlsx([snap], xlsx_path, pdf_files_by_email={snap.email: pdf_files})

        entry: dict[str, Any] = {
            "dir": acc_dir, "xlsx": xlsx_path, "json": None,
            "invoices": len(pdf_files),
        }

        if with_raw:
            json_path = acc_dir / "raw.json"
            json_path.write_text(
                json.dumps(asdict(snap), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            entry["json"] = json_path

        with _lock:
            summary[snap.email] = entry

        # 只有 with_detail_xlsx 时才需要再发 done（invoice 下载阶段已发过）
        if with_detail_xlsx:
            inv_count = len(pdf_files)
            _cb(snap.email, "done",
                f"账单 {inv_count} 份" if inv_count else "")

    if with_detail_xlsx or with_raw:
        max_workers = min(len(snapshots), 8)
        with _cf.ThreadPoolExecutor(max_workers=max_workers) as pool:
            futs = {pool.submit(_process_one, snap): snap for snap in snapshots}
            for fut in _cf.as_completed(futs):
                try:
                    fut.result()
                except Exception as e:
                    snap = futs[fut]
                    log.warning(f"[{snap.email}] export 阶段异常: {e}")
    else:
        # 仅补全 summary 字典（无需线程池）
        for snap in snapshots:
            acc_dir = out_root / _safe_filename(snap.email)
            acc_dir.mkdir(parents=True, exist_ok=True)
            pdf_files = all_pdf_files.get(snap.email, {})
            summary[snap.email] = {
                "dir": acc_dir, "xlsx": None, "json": None,
                "invoices": len(pdf_files),
            }

    # ── Phase 3：汇总 ──────────────────────────────────────────────
    if with_summary and snapshots:
        _cb("", "summary", "生成汇总报表...")
        if with_full_summary_xlsx:
            export_xlsx(snapshots, out_root / "_summary.xlsx",
                        pdf_files_by_email=all_pdf_files)
        export_token_summary_xlsx(
            snapshots,
            out_root / "汇总.xlsx",
            start_date=start_date,
            end_date=end_date,
        )
        _cb("", "zip_ready", "打包 ZIP 就绪")

    return summary
