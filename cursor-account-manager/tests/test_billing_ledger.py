"""账期净支出解析与汇总单测。"""

from decimal import Decimal

from cam.billing_ledger import (
    BillingListRow,
    export_billing_ledger_workbook,
    parse_amount_usd,
    parse_refund_amount_from_status,
    rows_from_billing_page_items,
    summarize_account_ledger,
)
from cam.exporter import (
    _billing_empty_from_refresh_state,
    _filter_billing_items_by_month,
)


def test_parse_amount_usd():
    assert parse_amount_usd("63.96 USD") == Decimal("63.96")
    assert parse_amount_usd("21.32 USD") == Decimal("21.32")


def test_parse_refunded_with_amount_in_status():
    status, refund_amt = parse_refund_amount_from_status("Refunded (12.07 USD)")
    assert status == "refunded"
    assert refund_amt == Decimal("12.07")


def test_net_spend_april_2026_screenshot_sample():
    """63.96 + 21.32 - 12.07 = 73.21"""
    rows = [
        BillingListRow(
            email="a@x.com",
            feishu_email="",
            date_text="2026年4月14日",
            description_text="",
            billing_month="2026-04",
            status_raw="Paid",
            status="paid",
            list_amount_usd=Decimal("63.96"),
            list_amount_raw="63.96 USD",
            refund_in_status_usd=None,
            invoice_url="https://invoice.stripe.com/i/1",
        ),
        BillingListRow(
            email="a@x.com",
            feishu_email="",
            date_text="2026年4月02日",
            description_text="",
            billing_month="2026-04",
            status_raw="Refunded (12.07 USD)",
            status="refunded",
            list_amount_usd=Decimal("21.32"),
            list_amount_raw="21.32 USD",
            refund_in_status_usd=Decimal("12.07"),
            invoice_url="https://invoice.stripe.com/i/2",
        ),
    ]
    s = summarize_account_ledger(
        rows, email="a@x.com", feishu_email="", billing_month="2026-04",
    )
    assert s.amount_total_usd == Decimal("85.28")
    assert s.refund_total_usd == Decimal("12.07")
    assert s.net_spend_usd == Decimal("73.21")


def test_rows_from_billing_page_items_filters_month():
    items = [
        {
            "url": "https://invoice.stripe.com/i/1",
            "date": "2026年4月14日",
            "description": "",
            "status": "Paid",
            "amountText": "63.96 USD",
        },
        {
            "url": "https://invoice.stripe.com/i/2",
            "date": "2026年3月01日",
            "description": "",
            "status": "Paid",
            "amountText": "10.00 USD",
        },
    ]
    rows = rows_from_billing_page_items(
        items, email="a@x.com", feishu_email="", selected_month="2026-04",
    )
    assert len(rows) == 1
    assert rows[0].list_amount_usd == Decimal("63.96")


def test_billing_empty_from_refresh_state_confirmed():
    payload = {"value": "2026-04", "labels": ["2026年4月"]}
    assert _billing_empty_from_refresh_state(
        {
            "ready": True,
            "selectedIndicator": True,
            "rowDates": [],
            "targetRowDates": [],
            "staleRowDates": [],
        },
        payload,
    )


def test_billing_empty_not_confirmed_while_loading():
    """仍在加载时不能判无账单。"""
    payload = {"value": "2026-04", "labels": ["2026年4月"]}
    assert not _billing_empty_from_refresh_state(
        {
            "ready": False,
            "selectedIndicator": True,
            "rowDates": [],
            "targetRowDates": [],
            "staleRowDates": [],
        },
        payload,
        loading=True,
    )


def test_billing_empty_confirmed_when_filter_set_but_stale_rows():
    """筛选已是目标月、仅残留旧月行 → 该月无账单，不重试。"""
    payload = {"value": "2026-05", "labels": ["2026年5月"]}
    assert _billing_empty_from_refresh_state(
        {
            "ready": False,
            "selectedIndicator": True,
            "rowDates": ["2026年4月21日"],
            "targetRowDates": [],
            "staleRowDates": ["2026年4月21日"],
        },
        payload,
        loading=False,
    )


def test_billing_empty_from_refresh_state_not_confirmed_when_stale_and_filter_wrong():
    payload = {"value": "2026-04", "labels": ["2026年4月"]}
    assert not _billing_empty_from_refresh_state(
        {
            "ready": False,
            "selectedIndicator": False,
            "rowDates": ["2026年3月14日"],
            "targetRowDates": [],
            "staleRowDates": ["2026年3月14日"],
        },
        payload,
    )


def test_filter_billing_items_by_month():
    items = [
        {"date": "2026年4月21日", "url": "https://invoice.stripe.com/i/1"},
        {"date": "2026年5月10日", "url": "https://invoice.stripe.com/i/2"},
    ]
    out = _filter_billing_items_by_month(items, "2026-05")
    assert len(out) == 1
    assert out[0]["date"] == "2026年5月10日"


def test_export_billing_ledger_workbook(tmp_path):
    from openpyxl import load_workbook

    summaries = [
        summarize_account_ledger([], email="a@x.com", feishu_email="", billing_month="2026-04"),
    ]
    path = export_billing_ledger_workbook(
        tmp_path / "ledger.xlsx", summaries, [],
    )
    wb = load_workbook(path)
    assert "账期净支出汇总" in wb.sheetnames
    assert "账单原始明细" in wb.sheetnames
    assert wb["账期净支出汇总"].max_row >= 2
